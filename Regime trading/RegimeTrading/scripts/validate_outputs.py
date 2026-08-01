from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from RegimeTrading.core.paths import DATA_DIR, INTRADAY_DB, POWERBI_WORKBOOK, legacy_output_path
from RegimeTrading.scripts.research_regime_aware_gap_recovery import (
    CANDIDATE_COLUMNS,
    DAILY_COLUMNS,
    SUMMARY_COLUMNS,
    TRADE_COLUMNS,
)
from RegimeTrading.scripts.step7_regime_feature_foundation import (
    AUDIT_COLUMNS as REGIME_FEATURE_AUDIT_COLUMNS,
    COMPLETENESS_COLUMNS as REGIME_FEATURE_COMPLETENESS_COLUMNS,
    DAILY_FEATURE_COLUMNS as REGIME_DAILY_FEATURE_COLUMNS,
    DEFINITION_COLUMNS as REGIME_FEATURE_DEFINITION_COLUMNS,
    SUMMARY_COLUMNS as REGIME_FEATURE_SUMMARY_COLUMNS,
)
from RegimeTrading.scripts.step7b_v1_regime_timing_comparison import (
    CANDIDATE_COLUMNS as REGIME_TIMING_CANDIDATE_COLUMNS,
    DAILY_COLUMNS as REGIME_TIMING_DAILY_COLUMNS,
    SUMMARY_COLUMNS as REGIME_TIMING_SUMMARY_COLUMNS,
    TRADE_COLUMNS as REGIME_TIMING_TRADE_COLUMNS,
)
from RegimeTrading.scripts.step8_provisional_regime_taxonomy import (
    DAILY_COLUMNS as REGIME_TAXONOMY_DAILY_COLUMNS,
    DEFINITION_COLUMNS as REGIME_TAXONOMY_DEFINITION_COLUMNS,
    DISTRIBUTION_COLUMNS as REGIME_TAXONOMY_DISTRIBUTION_COLUMNS,
    SUMMARY_COLUMNS as REGIME_TAXONOMY_SUMMARY_COLUMNS,
    TRANSITION_COLUMNS as REGIME_TAXONOMY_TRANSITION_COLUMNS,
)
from RegimeTrading.scripts.step9_playbook_specifications import (
    COVERAGE_COLUMNS as REGIME_PLAYBOOK_COVERAGE_COLUMNS,
    REGISTRY_COLUMNS as REGIME_PLAYBOOK_REGISTRY_COLUMNS,
    REQUIREMENT_COLUMNS as REGIME_PLAYBOOK_REQUIREMENT_COLUMNS,
    SUMMARY_COLUMNS as REGIME_PLAYBOOK_SUMMARY_COLUMNS,
)
from RegimeTrading.scripts.step9b_baseline_trade_generation import (
    AUDIT_COLUMNS as REGIME_BASELINE_AUDIT_COLUMNS,
    CANDIDATE_COLUMNS as REGIME_BASELINE_CANDIDATE_COLUMNS,
    LEG_COLUMNS as REGIME_BASELINE_LEG_COLUMNS,
    PERFORMANCE_COLUMNS as REGIME_BASELINE_PERFORMANCE_COLUMNS,
    SESSION_COLUMNS as REGIME_BASELINE_SESSION_COLUMNS,
    SUMMARY_COLUMNS as REGIME_BASELINE_SUMMARY_COLUMNS,
    TRADE_COLUMNS as REGIME_BASELINE_TRADE_COLUMNS,
)
from RegimeTrading.scripts.step9c_playbook_loss_diagnostics import (
    COST_SCENARIO_COLUMNS as REGIME_DIAGNOSTIC_COST_COLUMNS,
    LEAVE_ONE_DAY_OUT_COLUMNS as REGIME_DIAGNOSTIC_LOO_COLUMNS,
    PAIR_DIRECTION_COLUMNS as REGIME_DIAGNOSTIC_PAIR_COLUMNS,
    PLAYBOOK_DIAGNOSTIC_COLUMNS as REGIME_DIAGNOSTIC_PLAYBOOK_COLUMNS,
    SLICE_COLUMNS as REGIME_DIAGNOSTIC_SLICE_COLUMNS,
    SUMMARY_COLUMNS as REGIME_DIAGNOSTIC_SUMMARY_COLUMNS,
    TARGET_SCENARIO_COLUMNS as REGIME_DIAGNOSTIC_TARGET_COLUMNS,
    TRADE_DIAGNOSTIC_COLUMNS as REGIME_DIAGNOSTIC_TRADE_COLUMNS,
)
from RegimeTrading.scripts.step9d_regime_strategy_challenger_matrix import (
    AUDIT_COLUMNS as REGIME_CHALLENGER_AUDIT_COLUMNS,
    CANDIDATE_COLUMNS as REGIME_CHALLENGER_CANDIDATE_COLUMNS,
    LEG_COLUMNS as REGIME_CHALLENGER_LEG_COLUMNS,
    PERFORMANCE_COLUMNS as REGIME_CHALLENGER_PERFORMANCE_COLUMNS,
    RANKING_COLUMNS as REGIME_CHALLENGER_RANKING_COLUMNS,
    REGISTRY_COLUMNS as REGIME_CHALLENGER_REGISTRY_COLUMNS,
    SESSION_COLUMNS as REGIME_CHALLENGER_SESSION_COLUMNS,
    SUMMARY_COLUMNS as REGIME_CHALLENGER_SUMMARY_COLUMNS,
    TRADE_COLUMNS as REGIME_CHALLENGER_TRADE_COLUMNS,
)
from RegimeTrading.scripts.step9e_instrument_sector_taxonomy import (
    AUDIT_COLUMNS as INSTRUMENT_TAXONOMY_AUDIT_COLUMNS,
    CHARACTERISTIC_COLUMNS as INSTRUMENT_CHARACTERISTIC_COLUMNS,
    COMPLETENESS_COLUMNS as INSTRUMENT_TAXONOMY_COMPLETENESS_COLUMNS,
    CONSTRAINT_COLUMNS as INSTRUMENT_CONSTRAINT_COLUMNS,
    DEFINITION_COLUMNS as INSTRUMENT_DEFINITION_COLUMNS,
    GROUP_STATE_COLUMNS as INSTRUMENT_GROUP_STATE_COLUMNS,
    STATIC_COLUMNS as INSTRUMENT_STATIC_COLUMNS,
    SUMMARY_COLUMNS as INSTRUMENT_TAXONOMY_SUMMARY_COLUMNS,
)
from RegimeTrading.scripts.step9f_sector_ticker_strategy_experiments import (
    DIMENSION_AUDIT_COLUMNS as REGIME_SECTOR_DIMENSION_AUDIT_COLUMNS,
    EXCLUSION_COLUMNS as REGIME_SECTOR_EXCLUSION_COLUMNS,
    LEG_CONTEXT_COLUMNS as REGIME_SECTOR_LEG_CONTEXT_COLUMNS,
    PAIR_PERFORMANCE_COLUMNS as REGIME_SECTOR_PAIR_PERFORMANCE_COLUMNS,
    RANKING_COLUMNS as REGIME_SECTOR_RANKING_COLUMNS,
    SEGMENT_PERFORMANCE_COLUMNS as REGIME_SECTOR_SEGMENT_PERFORMANCE_COLUMNS,
    STATE_AUDIT_COLUMNS as REGIME_SECTOR_STATE_AUDIT_COLUMNS,
    SUMMARY_COLUMNS as REGIME_SECTOR_SUMMARY_COLUMNS,
    TRADE_CONTEXT_COLUMNS as REGIME_SECTOR_TRADE_CONTEXT_COLUMNS,
)
from RegimeTrading.scripts.step9g_state_filtered_contract_experiments import (
    AUDIT_COLUMNS as STATE_FILTER_AUDIT_COLUMNS,
    CANDIDATE_COLUMNS as STATE_FILTER_CANDIDATE_COLUMNS,
    COMPARISON_COLUMNS as STATE_FILTER_COMPARISON_COLUMNS,
    LEG_COLUMNS as STATE_FILTER_LEG_COLUMNS,
    MULTIPLE_TESTING_COLUMNS as STATE_FILTER_MULTIPLE_TESTING_COLUMNS,
    PERFORMANCE_COLUMNS as STATE_FILTER_PERFORMANCE_COLUMNS,
    REGISTRY_COLUMNS as STATE_FILTER_REGISTRY_COLUMNS,
    ROBUSTNESS_COLUMNS as STATE_FILTER_ROBUSTNESS_COLUMNS,
    SESSION_COLUMNS as STATE_FILTER_SESSION_COLUMNS,
    SUMMARY_COLUMNS as STATE_FILTER_SUMMARY_COLUMNS,
    TRADE_COLUMNS as STATE_FILTER_TRADE_COLUMNS,
)
from RegimeTrading.scripts.v1_validation_concentration import (
    CONTRIBUTION_COLUMNS,
    LEAVE_ONE_OUT_COLUMNS,
    SCENARIO_COLUMNS as CONCENTRATION_SCENARIO_COLUMNS,
    SUMMARY_COLUMNS as CONCENTRATION_SUMMARY_COLUMNS,
)
from RegimeTrading.scripts.v1_validation_exposure_reconciliation import (
    RECONCILIATION_COLUMNS as EXPOSURE_RECONCILIATION_COLUMNS,
)
from RegimeTrading.scripts.v1_validation_exposure_efficiency import (
    DAILY_COLUMNS as EXPOSURE_DAILY_COLUMNS,
    INTERVAL_DETAIL_COLUMNS as EXPOSURE_INTERVAL_COLUMNS,
    POSITION_DETAIL_COLUMNS as EXPOSURE_POSITION_COLUMNS,
    SIZING_COLUMNS as EXPOSURE_SIZING_COLUMNS,
    SUMMARY_COLUMNS as EXPOSURE_SUMMARY_COLUMNS,
)
from RegimeTrading.scripts.v1_validation_execution_stress import (
    COST_CURVE_COLUMNS as EXECUTION_COST_CURVE_COLUMNS,
    SCENARIO_COLUMNS as EXECUTION_SCENARIO_COLUMNS,
    SUMMARY_COLUMNS as EXECUTION_SUMMARY_COLUMNS,
    TRADE_DETAIL_COLUMNS as EXECUTION_TRADE_DETAIL_COLUMNS,
)
from RegimeTrading.scripts.v1_validation_parameter_robustness import (
    RECONCILIATION_COLUMNS as PARAMETER_RECONCILIATION_COLUMNS,
    SCENARIO_COLUMNS as PARAMETER_SCENARIO_COLUMNS,
    SENSITIVITY_COLUMNS as PARAMETER_SENSITIVITY_COLUMNS,
    SUMMARY_COLUMNS as PARAMETER_SUMMARY_COLUMNS,
)
from RegimeTrading.scripts.v1_validation_provider_quality import (
    DAILY_COLUMNS as PROVIDER_DAILY_COLUMNS,
    MISMATCH_COLUMNS as PROVIDER_MISMATCH_COLUMNS,
    SESSION_COLUMNS as PROVIDER_SESSION_COLUMNS,
    SUMMARY_COLUMNS as PROVIDER_SUMMARY_COLUMNS,
)
from RegimeTrading.scripts.v1_validation_portfolio import (
    DAILY_COLUMNS as PORTFOLIO_DAILY_COLUMNS,
    EQUITY_COLUMNS as PORTFOLIO_EQUITY_COLUMNS,
    LEDGER_COLUMNS as PORTFOLIO_LEDGER_COLUMNS,
    SUMMARY_COLUMNS as PORTFOLIO_SUMMARY_COLUMNS,
)


EXPECTED = {
    "regime_gap_recovery_summary.csv": SUMMARY_COLUMNS,
    "regime_gap_recovery_trades.csv": TRADE_COLUMNS,
    "regime_gap_recovery_daily.csv": DAILY_COLUMNS,
    "regime_gap_recovery_latest.csv": CANDIDATE_COLUMNS,
    "regime_gap_recovery_candidates.csv": CANDIDATE_COLUMNS,
    "regime_gap_recovery_forward_summary.csv": SUMMARY_COLUMNS,
    "regime_gap_recovery_forward_trades.csv": TRADE_COLUMNS,
    "regime_gap_recovery_forward_daily.csv": DAILY_COLUMNS,
    "regime_gap_recovery_forward_candidates.csv": CANDIDATE_COLUMNS,
    "v1_validation_portfolio_summary.csv": PORTFOLIO_SUMMARY_COLUMNS,
    "v1_validation_portfolio_trade_ledger.csv": PORTFOLIO_LEDGER_COLUMNS,
    "v1_validation_portfolio_equity_curve.csv": PORTFOLIO_EQUITY_COLUMNS,
    "v1_validation_portfolio_daily.csv": PORTFOLIO_DAILY_COLUMNS,
    "v1_validation_concentration_summary.csv": CONCENTRATION_SUMMARY_COLUMNS,
    "v1_validation_concentration_scenarios.csv": CONCENTRATION_SCENARIO_COLUMNS,
    "v1_validation_contribution_detail.csv": CONTRIBUTION_COLUMNS,
    "v1_validation_leave_one_out.csv": LEAVE_ONE_OUT_COLUMNS,
    "v1_validation_execution_stress_summary.csv": EXECUTION_SUMMARY_COLUMNS,
    "v1_validation_execution_stress_scenarios.csv": EXECUTION_SCENARIO_COLUMNS,
    "v1_validation_execution_stress_trade_detail.csv": EXECUTION_TRADE_DETAIL_COLUMNS,
    "v1_validation_execution_cost_curve.csv": EXECUTION_COST_CURVE_COLUMNS,
    "v1_validation_parameter_robustness_summary.csv": PARAMETER_SUMMARY_COLUMNS,
    "v1_validation_parameter_robustness_scenarios.csv": PARAMETER_SCENARIO_COLUMNS,
    "v1_validation_parameter_sensitivity.csv": PARAMETER_SENSITIVITY_COLUMNS,
    "v1_validation_parameter_baseline_reconciliation.csv": PARAMETER_RECONCILIATION_COLUMNS,
    "v1_validation_provider_quality_summary.csv": PROVIDER_SUMMARY_COLUMNS,
    "v1_validation_provider_session_detail.csv": PROVIDER_SESSION_COLUMNS,
    "v1_validation_provider_daily_summary.csv": PROVIDER_DAILY_COLUMNS,
    "v1_validation_provider_mismatch_detail.csv": PROVIDER_MISMATCH_COLUMNS,
    "v1_validation_exposure_efficiency_summary.csv": EXPOSURE_SUMMARY_COLUMNS,
    "v1_validation_exposure_position_detail.csv": EXPOSURE_POSITION_COLUMNS,
    "v1_validation_exposure_interval_detail.csv": EXPOSURE_INTERVAL_COLUMNS,
    "v1_validation_exposure_daily.csv": EXPOSURE_DAILY_COLUMNS,
    "v1_validation_position_size_scenarios.csv": EXPOSURE_SIZING_COLUMNS,
    "v1_validation_exposure_reconciliation.csv": EXPOSURE_RECONCILIATION_COLUMNS,
    "regime_feature_foundation_summary.csv": REGIME_FEATURE_SUMMARY_COLUMNS,
    "regime_daily_features.csv": REGIME_DAILY_FEATURE_COLUMNS,
    "regime_feature_definitions.csv": REGIME_FEATURE_DEFINITION_COLUMNS,
    "regime_feature_completeness.csv": REGIME_FEATURE_COMPLETENESS_COLUMNS,
    "regime_point_in_time_audit.csv": REGIME_FEATURE_AUDIT_COLUMNS,
    "regime_v1_timing_comparison_summary.csv": REGIME_TIMING_SUMMARY_COLUMNS,
    "regime_v1_timing_comparison_daily.csv": REGIME_TIMING_DAILY_COLUMNS,
    "regime_v1_timing_comparison_candidates.csv": REGIME_TIMING_CANDIDATE_COLUMNS,
    "regime_v1_timing_comparison_trades.csv": REGIME_TIMING_TRADE_COLUMNS,
    "regime_taxonomy_summary.csv": REGIME_TAXONOMY_SUMMARY_COLUMNS,
    "regime_daily_taxonomy.csv": REGIME_TAXONOMY_DAILY_COLUMNS,
    "regime_taxonomy_definitions.csv": REGIME_TAXONOMY_DEFINITION_COLUMNS,
    "regime_taxonomy_distribution.csv": REGIME_TAXONOMY_DISTRIBUTION_COLUMNS,
    "regime_taxonomy_transitions.csv": REGIME_TAXONOMY_TRANSITION_COLUMNS,
    "regime_playbook_specification_summary.csv": REGIME_PLAYBOOK_SUMMARY_COLUMNS,
    "regime_playbook_registry.csv": REGIME_PLAYBOOK_REGISTRY_COLUMNS,
    "regime_playbook_data_requirements.csv": REGIME_PLAYBOOK_REQUIREMENT_COLUMNS,
    "regime_playbook_session_coverage.csv": REGIME_PLAYBOOK_COVERAGE_COLUMNS,
    "regime_playbook_baseline_summary.csv": REGIME_BASELINE_SUMMARY_COLUMNS,
    "regime_playbook_baseline_sessions.csv": REGIME_BASELINE_SESSION_COLUMNS,
    "regime_playbook_baseline_candidates.csv": REGIME_BASELINE_CANDIDATE_COLUMNS,
    "regime_playbook_baseline_trades.csv": REGIME_BASELINE_TRADE_COLUMNS,
    "regime_playbook_baseline_trade_legs.csv": REGIME_BASELINE_LEG_COLUMNS,
    "regime_playbook_baseline_performance.csv": REGIME_BASELINE_PERFORMANCE_COLUMNS,
    "regime_playbook_baseline_audit.csv": REGIME_BASELINE_AUDIT_COLUMNS,
    "regime_playbook_diagnostic_summary.csv": REGIME_DIAGNOSTIC_SUMMARY_COLUMNS,
    "regime_playbook_trade_diagnostics.csv": REGIME_DIAGNOSTIC_TRADE_COLUMNS,
    "regime_playbook_diagnostics.csv": REGIME_DIAGNOSTIC_PLAYBOOK_COLUMNS,
    "regime_playbook_diagnostic_slices.csv": REGIME_DIAGNOSTIC_SLICE_COLUMNS,
    "regime_playbook_target_scenarios.csv": REGIME_DIAGNOSTIC_TARGET_COLUMNS,
    "regime_playbook_cost_scenarios.csv": REGIME_DIAGNOSTIC_COST_COLUMNS,
    "regime_playbook_leave_one_day_out.csv": REGIME_DIAGNOSTIC_LOO_COLUMNS,
    "regime_playbook_pair_direction_controls.csv": REGIME_DIAGNOSTIC_PAIR_COLUMNS,
    "regime_challenger_matrix_summary.csv": REGIME_CHALLENGER_SUMMARY_COLUMNS,
    "regime_challenger_registry.csv": REGIME_CHALLENGER_REGISTRY_COLUMNS,
    "regime_challenger_candidates.csv": REGIME_CHALLENGER_CANDIDATE_COLUMNS,
    "regime_challenger_trades.csv": REGIME_CHALLENGER_TRADE_COLUMNS,
    "regime_challenger_trade_legs.csv": REGIME_CHALLENGER_LEG_COLUMNS,
    "regime_challenger_performance.csv": REGIME_CHALLENGER_PERFORMANCE_COLUMNS,
    "regime_challenger_rankings.csv": REGIME_CHALLENGER_RANKING_COLUMNS,
    "regime_challenger_session_coverage.csv": REGIME_CHALLENGER_SESSION_COLUMNS,
    "regime_challenger_audit.csv": REGIME_CHALLENGER_AUDIT_COLUMNS,
    "instrument_taxonomy_summary.csv": INSTRUMENT_TAXONOMY_SUMMARY_COLUMNS,
    "instrument_static_taxonomy.csv": INSTRUMENT_STATIC_COLUMNS,
    "instrument_characteristic_definitions.csv": INSTRUMENT_DEFINITION_COLUMNS,
    "instrument_point_in_time_characteristics.csv": INSTRUMENT_CHARACTERISTIC_COLUMNS,
    "instrument_group_daily_state.csv": INSTRUMENT_GROUP_STATE_COLUMNS,
    "instrument_taxonomy_completeness.csv": INSTRUMENT_TAXONOMY_COMPLETENESS_COLUMNS,
    "instrument_relationship_constraints.csv": INSTRUMENT_CONSTRAINT_COLUMNS,
    "instrument_taxonomy_audit.csv": INSTRUMENT_TAXONOMY_AUDIT_COLUMNS,
    "regime_sector_strategy_summary.csv": REGIME_SECTOR_SUMMARY_COLUMNS,
    "regime_sector_strategy_trade_context.csv": REGIME_SECTOR_TRADE_CONTEXT_COLUMNS,
    "regime_sector_strategy_leg_context.csv": REGIME_SECTOR_LEG_CONTEXT_COLUMNS,
    "regime_sector_strategy_segment_performance.csv": REGIME_SECTOR_SEGMENT_PERFORMANCE_COLUMNS,
    "regime_sector_strategy_pair_performance.csv": REGIME_SECTOR_PAIR_PERFORMANCE_COLUMNS,
    "regime_sector_strategy_exclusion_robustness.csv": REGIME_SECTOR_EXCLUSION_COLUMNS,
    "regime_sector_strategy_dimension_audit.csv": REGIME_SECTOR_DIMENSION_AUDIT_COLUMNS,
    "regime_sector_strategy_state_audit.csv": REGIME_SECTOR_STATE_AUDIT_COLUMNS,
    "regime_sector_strategy_rankings.csv": REGIME_SECTOR_RANKING_COLUMNS,
    "regime_state_filtered_summary.csv": STATE_FILTER_SUMMARY_COLUMNS,
    "regime_state_filtered_contract_registry.csv": STATE_FILTER_REGISTRY_COLUMNS,
    "regime_state_filtered_session_coverage.csv": STATE_FILTER_SESSION_COLUMNS,
    "regime_state_filtered_candidates.csv": STATE_FILTER_CANDIDATE_COLUMNS,
    "regime_state_filtered_trades.csv": STATE_FILTER_TRADE_COLUMNS,
    "regime_state_filtered_trade_legs.csv": STATE_FILTER_LEG_COLUMNS,
    "regime_state_filtered_performance.csv": STATE_FILTER_PERFORMANCE_COLUMNS,
    "regime_state_filtered_same_cohort_comparisons.csv": STATE_FILTER_COMPARISON_COLUMNS,
    "regime_state_filtered_robustness.csv": STATE_FILTER_ROBUSTNESS_COLUMNS,
    "regime_state_filtered_multiple_testing.csv": STATE_FILTER_MULTIPLE_TESTING_COLUMNS,
    "regime_state_filtered_audit.csv": STATE_FILTER_AUDIT_COLUMNS,
}


def main() -> None:
    print("\n=== VALIDATE REGIME TRADING OUTPUTS ===")
    errors: list[str] = []

    if not INTRADAY_DB.exists():
        errors.append(f"Missing local database: {INTRADAY_DB}")

    for filename, expected_columns in EXPECTED.items():
        path = legacy_output_path(filename)
        if not path.exists():
            errors.append(f"Missing output: {path}")
            continue

        dataframe = pd.read_csv(path)
        missing_columns = [column for column in expected_columns if column not in dataframe]
        if missing_columns:
            errors.append(
                f"{filename} missing columns: {', '.join(missing_columns)}"
            )
        else:
            print(f"OK: {filename} ({len(dataframe)} rows)")

    if not POWERBI_WORKBOOK.exists():
        errors.append(f"Missing Power BI workbook: {POWERBI_WORKBOOK}")
    else:
        workbook = load_workbook(POWERBI_WORKBOOK, read_only=True)
        print(f"OK: {POWERBI_WORKBOOK.name} ({len(workbook.sheetnames)} sheets)")
        workbook.close()

    if errors:
        print("\nValidation failed:")
        for error in errors:
            print(f"- {error}")
        sys.exit(1)

    print("\nAll isolated research outputs are valid.")


if __name__ == "__main__":
    main()
