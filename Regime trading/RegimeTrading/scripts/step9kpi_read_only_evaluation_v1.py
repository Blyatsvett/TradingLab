from __future__ import annotations

import argparse
import hashlib
import itertools
import json
import math
import os
import sqlite3
import tempfile
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from RegimeTrading.core.paths import KPI_OUTPUT_DIR


SPECIFICATION_ID = "STEP9KPI_READ_ONLY_EVALUATION_V1"
SCHEMA_ID = "STEP9_KPI_OUTPUT_SCHEMA_V1"
IMPLEMENTATION_ID = "STEP9KPI_READ_ONLY_EVALUATION_V1_1_2026_07_29"
STATUS = "READ_ONLY_KPI_EVALUATION_NOT_SELECTOR_NOT_ROUTER_ACTIVE"

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG = PROJECT_ROOT / "config" / "step9kpi_read_only_evaluation_v1.json"
DEFAULT_SCHEMA = PROJECT_ROOT / "config" / "step9kpi_output_schema_v1.json"
DEFAULT_OUTPUT_DIR = KPI_OUTPUT_DIR
DEFAULT_WORKBOOK = DEFAULT_OUTPUT_DIR / "powerbi_step9_kpi_monitor.xlsx"


class Step9KpiError(RuntimeError):
    pass


class SourceContractError(Step9KpiError):
    pass


class OutputContractError(Step9KpiError):
    pass


@dataclass(frozen=True)
class SourceSpec:
    source_id: str
    path: Path
    required: bool


@dataclass
class BuildResult:
    tables: dict[str, pd.DataFrame]
    output_files: list[Path]
    source_hashes_before: dict[str, str]
    source_hashes_after: dict[str, str]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_hash(*parts: Any) -> str:
    text = "|".join("" if part is None else str(part) for part in parts)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _sqlite_uri(path: Path) -> str:
    absolute = Path(path).resolve().as_posix()
    return f"file:{quote(absolute, safe='/:')}?mode=ro"


def _connect_ro(path: Path) -> sqlite3.Connection:
    if not Path(path).exists():
        raise SourceContractError(f"Source database does not exist: {path}")
    connection = sqlite3.connect(_sqlite_uri(path), uri=True)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA query_only=ON")
    return connection


def _table_names(connection: sqlite3.Connection) -> set[str]:
    return {
        str(row[0])
        for row in connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()
    }


def _read_table(path: Path, table: str, where: str | None = None, params: tuple[Any, ...] = ()) -> pd.DataFrame:
    with _connect_ro(path) as connection:
        if table not in _table_names(connection):
            return pd.DataFrame()
        query = f'SELECT * FROM "{table}"'
        if where:
            query += f" WHERE {where}"
        return pd.read_sql_query(query, connection, params=params)


def _read_json(path: Path, default: Any = None) -> Any:
    if not Path(path).exists():
        return default
    return json.loads(Path(path).read_text(encoding="utf-8"))


def _read_config(path: Path) -> dict[str, Any]:
    payload = _read_json(path)
    if not isinstance(payload, dict):
        raise SourceContractError(f"Invalid KPI config: {path}")
    if payload.get("specification_id") != SPECIFICATION_ID:
        raise SourceContractError(
            f"Expected specification_id={SPECIFICATION_ID}; found {payload.get('specification_id')!r}."
        )
    if payload.get("router_active") is not False or payload.get("orders_enabled") is not False:
        raise SourceContractError("KPI config must keep router and orders disabled.")
    return payload


def _read_schema(path: Path) -> dict[str, Any]:
    payload = _read_json(path)
    if not isinstance(payload, dict) or payload.get("schema_id") != SCHEMA_ID:
        raise OutputContractError(f"Invalid KPI schema: {path}")
    if "tables" not in payload:
        raise OutputContractError("KPI schema has no tables.")
    return payload


def _safe_float(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return None if not math.isfinite(number) else number


def _safe_int(value: Any, default: int = 0) -> int:
    number = _safe_float(value)
    return default if number is None else int(number)


def _bool(value: Any) -> bool:
    if value is None or pd.isna(value):
        return False
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y"}
    return bool(value)


def _evidence_status(prospective_status: Any) -> str:
    text = str(prospective_status or "").upper()
    if "SIMULATED" in text or "MOCK" in text:
        return "MOCK_REHEARSAL"
    if "PROSPECTIVE_CONFIRMATORY" in text and "NOT" not in text:
        return "PROSPECTIVE_CONFIRMATORY"
    if "HISTORICAL" in text or "RETROSPECTIVE" in text:
        return "HISTORICAL_RETROSPECTIVE"
    return "PROSPECTIVE_EXCLUDED"


def _standardize_return(gross_return: Any, notional: float, cost_rate: float) -> tuple[float | None, float | None]:
    value = _safe_float(gross_return)
    if value is None:
        return None, None
    cost = notional * cost_rate
    return cost, value * notional - cost


def _normalize_outcome_status(status: Any, gross_return: Any = None) -> str:
    text = str(status or "").upper()
    if "NO_COMPLETED_TRADE" in text or "NO_TRIGGER" in text or "NOT_TRIGGERED" in text:
        return "NO_TRIGGER"
    if _safe_float(gross_return) is not None or "COMPLETE" in text or "COMPLETED" in text:
        return "COMPLETE"
    if "CASH" in text:
        return "ZERO_CASH"
    return "INCOMPLETE"


def _datetime_or_none(value: Any) -> str | None:
    if value is None or pd.isna(value) or not str(value).strip():
        return None
    try:
        return pd.Timestamp(value).isoformat(sep=" ")
    except Exception:
        return str(value)


def _build_sources(project_root: Path, config: dict[str, Any]) -> dict[str, SourceSpec]:
    paths = config["source_paths"]
    required = {"price_db", "step9i", "step9l"}
    return {
        key: SourceSpec(key, project_root / relative, key in required)
        for key, relative in paths.items()
    }


def _existing_source_hashes(sources: dict[str, SourceSpec]) -> dict[str, str]:
    hashes: dict[str, str] = {}
    for source_id, spec in sources.items():
        if spec.path.exists():
            hashes[source_id] = _sha256(spec.path)
        elif spec.required:
            raise SourceContractError(f"Required source is missing: {spec.path}")
    return hashes


def _load_engine_batches(sources: dict[str, SourceSpec]) -> dict[str, pd.DataFrame]:
    mapping = {
        "step9i": ("shadow_decision_batches", "shadow_outcome_batches"),
        "step9l": ("shadow_decision_batches", "shadow_outcome_batches"),
        "step9s": ("step9s_assignments", "step9s_outcome_batches"),
        "step9r": ("selector_batches", None),
        "step9t": ("step9t_prospective_batches", "step9t_prospective_outcome_batches"),
        "step9u": ("step9u_prospective_assignment_batches", "step9u_prospective_outcome_batches"),
        "step9v": ("step9v_checkpoint_batches", "step9v_checkpoint_outcome_batches"),
    }
    output: dict[str, pd.DataFrame] = {}
    for source_id, (morning_table, eod_table) in mapping.items():
        source = sources[source_id].path
        if not source.exists():
            output[source_id] = pd.DataFrame()
            continue
        morning = _read_table(source, morning_table)
        if morning.empty:
            output[source_id] = morning
            continue
        if eod_table:
            eod = _read_table(source, eod_table)
            if not eod.empty and "session_date" in eod.columns:
                eod_cols = [c for c in eod.columns if c != "session_date"]
                eod = eod.sort_values(["session_date"] + (["checkpoint_time"] if "checkpoint_time" in eod.columns else [])).drop_duplicates("session_date", keep="last")
                morning = morning.merge(eod[["session_date"] + eod_cols], on="session_date", how="left", suffixes=("", "_eod"))
        output[source_id] = morning
    return output


def _session_evidence_map(batches: dict[str, pd.DataFrame]) -> dict[str, str]:
    priority = {
        "PROSPECTIVE_CONFIRMATORY": 4,
        "MOCK_REHEARSAL": 3,
        "HISTORICAL_RETROSPECTIVE": 2,
        "PROSPECTIVE_EXCLUDED": 1,
    }
    result: dict[str, str] = {}
    for frame in batches.values():
        if frame.empty or "session_date" not in frame.columns:
            continue
        for row in frame.to_dict("records"):
            status = _evidence_status(row.get("prospective_status"))
            session = str(row["session_date"])
            if priority[status] >= priority.get(result.get(session, ""), 0):
                result[session] = status
    return result


def _morning_regime_map(batches: dict[str, pd.DataFrame]) -> dict[str, str | None]:
    for key in ("step9l", "step9i", "step9t", "step9u"):
        frame = batches.get(key, pd.DataFrame())
        if frame.empty:
            continue
        regime_column = next((c for c in ("primary_regime", "source_regime", "morning_regime") if c in frame.columns), None)
        if regime_column:
            return {
                str(row["session_date"]): (None if pd.isna(row[regime_column]) else str(row[regime_column]))
                for row in frame.sort_values("session_date").drop_duplicates("session_date", keep="last").to_dict("records")
            }
    return {}


def _source_batches_by_session(batches: dict[str, pd.DataFrame]) -> dict[tuple[str, str], str | None]:
    result: dict[tuple[str, str], str | None] = {}
    candidates = {
        "step9i": "outcome_batch_id",
        "step9l": "outcome_batch_id",
        "step9s": "outcome_batch_id",
        "step9r": "batch_id",
        "step9t": "outcome_batch_id",
        "step9u": "outcome_batch_id",
        "step9v": "outcome_batch_id",
    }
    for source, frame in batches.items():
        if frame.empty:
            continue
        id_col = candidates[source] if candidates[source] in frame.columns else next((c for c in ("batch_id", "assignment_id", "assignment_batch_id", "checkpoint_batch_id") if c in frame.columns), None)
        for row in frame.sort_values("session_date").drop_duplicates("session_date", keep="last").to_dict("records"):
            result[(source, str(row["session_date"]))] = None if id_col is None or pd.isna(row.get(id_col)) else str(row.get(id_col))
    return result


def _load_universes(project_root: Path, config: dict[str, Any], sources: dict[str, SourceSpec]) -> tuple[list[str], list[str]]:
    q_schema = project_root / config["optional_registry_paths"]["step9q_schema"]
    payload = _read_json(q_schema, {}) or {}
    universe = payload.get("universe", {})
    regime = [str(x) for x in universe.get("regime_source_tickers", [])]
    tradable = [str(x) for x in universe.get("tradable_tickers", [])]
    broad = sorted(set(regime) | set(tradable))
    if regime and broad:
        return regime, broad

    if sources["step9t"].path.exists():
        tickers = _read_table(sources["step9t"].path, "step9t_prospective_ticker_archetypes")
        if not tickers.empty:
            regime = sorted(tickers.loc[tickers["universe_role"].eq("REGIME_SOURCE"), "ticker"].astype(str).unique())
            broad = sorted(tickers["ticker"].astype(str).unique())
    if not broad:
        prices = _read_table(sources["price_db"].path, "intraday_prices")
        broad = sorted(prices["ticker"].astype(str).unique())
    return regime, broad


def _strategy_membership_index(sources: dict[str, SourceSpec]) -> dict[tuple[str, str], set[str]]:
    memberships: dict[tuple[str, str], set[str]] = {}
    def add(frame: pd.DataFrame, source: str, ticker_col: str = "ticker") -> None:
        if frame.empty:
            return
        for row in frame.to_dict("records"):
            key = (str(row.get("session_date")), str(row.get(ticker_col)))
            memberships.setdefault(key, set()).add(source)
    if sources["step9s"].path.exists():
        add(_read_table(sources["step9s"].path, "step9s_natural_outcomes"), "STEP9S_NATURAL")
    if sources["step9r"].path.exists():
        add(_read_table(sources["step9r"].path, "selector_candidate_outcomes"), "STEP9R_CANDIDATE")
    if sources["step9u"].path.exists():
        add(_read_table(sources["step9u"].path, "step9u_prospective_candidate_outcomes"), "STEP9U_CANDIDATE")
    return memberships


def _strategy_outcomes(
    sources: dict[str, SourceSpec],
    evidence: dict[str, str],
    config: dict[str, Any],
) -> pd.DataFrame:
    notional = float(config["comparison_notional_sek"])
    cost_rate = float(config["comparison_round_trip_cost_rate"])
    rows: list[dict[str, Any]] = []
    membership_index = _strategy_membership_index(sources)

    # Authoritative Step 9L candidate outcomes.
    if sources["step9l"].path.exists():
        frame = _read_table(sources["step9l"].path, "shadow_outcomes")
        frame = frame.loc[frame.get("candidate_generated", 0).fillna(0).astype(int).eq(1)] if not frame.empty else frame
        for row in frame.to_dict("records"):
            session = str(row["session_date"])
            status = _normalize_outcome_status(row.get("outcome_status"), row.get("gross_return"))
            cost, standardized = _standardize_return(row.get("gross_return"), notional, cost_rate)
            if status == "NO_TRIGGER":
                cost, standardized = 0.0, 0.0
            variant = f"STEP9L::{row.get('contract_id')}"
            memberships = {"STEP9L"} | membership_index.get((session, str(row.get("ticker"))), set())
            rows.append(_outcome_row(
                session, evidence.get(session, "PROSPECTIVE_EXCLUDED"), str(row.get("ticker")), row.get("broad_sector"),
                variant, row.get("direction"), bool(row.get("morning_contract_eligible")), True, status,
                row.get("entry_time"), row.get("exit_time"), 750.0,
                row.get("risk_capped_net_pnl_sek"), row.get("gross_return"), cost, standardized,
                None, None, "step9l_v3_selected_strategy_shadow_ledger.db/shadow_outcomes",
                memberships,
            ))

    # Step 9T archetype 09:50-to-EOD outcomes.
    if sources["step9t"].path.exists():
        frame = _read_table(sources["step9t"].path, "step9t_prospective_ticker_outcomes")
        sectors = _read_table(sources["step9t"].path, "step9t_prospective_ticker_archetypes")
        sector_map = {(str(r["session_date"]), str(r["ticker"])): r.get("broad_sector") for r in sectors.to_dict("records")} if not sectors.empty else {}
        for row in frame.to_dict("records"):
            if str(row.get("direction", "")).upper() not in {"LONG", "SHORT"}:
                continue
            session = str(row["session_date"])
            status = _normalize_outcome_status(row.get("outcome_status"), row.get("session_close_return"))
            variant = f"STEP9T::{row.get('primary_archetype')}::0950_TO_EOD"
            memberships = {"STEP9T"} | membership_index.get((session, str(row.get("ticker"))), set())
            rows.append(_outcome_row(
                session, evidence.get(session, "PROSPECTIVE_EXCLUDED"), str(row.get("ticker")),
                sector_map.get((session, str(row.get("ticker")))), variant, row.get("direction"), True, False,
                status, row.get("entry_time"), row.get("exit_time"), notional,
                row.get("net_pnl_sek"), row.get("session_close_return"), row.get("cost_sek"), row.get("net_pnl_sek"),
                row.get("mfe_return"), row.get("mae_return"),
                "step9t_regime_transition_archetype_prospective_v1.db/step9t_prospective_ticker_outcomes",
                memberships,
            ))

    # Step 9S mandatory control outcomes.
    if sources["step9s"].path.exists():
        frame = _read_table(sources["step9s"].path, "step9s_coverage_outcomes")
        for row in frame.to_dict("records"):
            session = str(row["session_date"])
            ticker = str(row.get("ticker") or row.get("paired_ticker") or f"PAIR::{row.get('long_ticker')}::{row.get('short_ticker')}")
            variant = f"STEP9S_CONTROL::{row.get('coverage_control_id')}"
            status = _normalize_outcome_status("COMPLETE", row.get("gross_return"))
            standardized_cost, standardized = _standardize_return(row.get("gross_return"), notional, cost_rate)
            rows.append(_outcome_row(
                session, evidence.get(session, "PROSPECTIVE_EXCLUDED"), ticker, None, variant, row.get("direction"), True, True,
                status, row.get("entry_time"), row.get("exit_time"), row.get("notional_sek"), row.get("net_pnl_sek"),
                row.get("gross_return"), standardized_cost, standardized, None, None,
                "step9s_prospective_contingency_shadow_v1.db/step9s_coverage_outcomes", {"STEP9S_MANDATORY_CONTROL"},
            ))

    # Step 9V management variants. Deliberately not benchmark-comparison eligible.
    if sources["step9v"].path.exists():
        frame = _read_table(sources["step9v"].path, "step9v_selected_action_outcomes")
        reviews = _read_table(sources["step9v"].path, "step9v_selected_position_reviews")
        review_map = {str(r["review_id"]): r for r in reviews.to_dict("records")} if not reviews.empty else {}
        for row in frame.to_dict("records"):
            review = review_map.get(str(row.get("review_id")), {})
            session = str(row["session_date"])
            checkpoint = str(row.get("checkpoint_time"))
            base_rule = str(review.get("rule_id") or "UNKNOWN_RULE")
            for action, field in (
                ("KEEP", "hold_net_pnl_sek"),
                ("REDUCE", "reduce_net_pnl_sek"),
                ("EXIT", "exit_net_pnl_sek"),
                ("SWITCH", "switch_net_pnl_sek"),
            ):
                pnl = _safe_float(row.get(field))
                if pnl is None:
                    continue
                variant = f"STEP9V::{base_rule}::{checkpoint}::{action}"
                gross_return = (pnl + notional * cost_rate) / notional
                rows.append(_outcome_row(
                    session, evidence.get(session, "PROSPECTIVE_EXCLUDED"), str(row.get("ticker")), review.get("broad_sector"),
                    variant, row.get("original_direction") if action != "SWITCH" else "DYNAMIC", True, False, "COMPLETE",
                    row.get("morning_entry_label"), row.get("exit_label"), notional, pnl, gross_return,
                    notional * cost_rate, pnl, None, None,
                    "step9v_intraday_regime_transition_observer_v1.db/step9v_selected_action_outcomes", {"STEP9V"},
                ))

    if not rows:
        return pd.DataFrame()
    frame = pd.DataFrame(rows)
    # Strict source precedence and duplicate key. Distinct strategy variants remain separate.
    precedence = {
        "step9l_v3_selected_strategy_shadow_ledger.db/shadow_outcomes": 1,
        "step9t_regime_transition_archetype_prospective_v1.db/step9t_prospective_ticker_outcomes": 2,
        "step9s_prospective_contingency_shadow_v1.db/step9s_coverage_outcomes": 3,
        "step9v_intraday_regime_transition_observer_v1.db/step9v_selected_action_outcomes": 4,
    }
    frame["_precedence"] = frame["authoritative_source"].map(precedence).fillna(99)
    frame = frame.sort_values(["canonical_outcome_id", "_precedence"]).drop_duplicates("canonical_outcome_id", keep="first")
    return frame.drop(columns=["_precedence"]).reset_index(drop=True)


def _outcome_row(
    session_date: str,
    evidence_status: str,
    ticker: str,
    broad_sector: Any,
    strategy_variant_id: str,
    direction: Any,
    morning_available: bool,
    governance_feasible: bool,
    outcome_status: str,
    entry_time: Any,
    exit_time: Any,
    native_notional: Any,
    native_pnl: Any,
    gross_return: Any,
    standardized_cost: Any,
    standardized_pnl: Any,
    mfe: Any,
    mae: Any,
    source: str,
    memberships: Iterable[str],
) -> dict[str, Any]:
    entry = _datetime_or_none(entry_time)
    exit_ = _datetime_or_none(exit_time)
    canonical_id = _canonical_hash(session_date, ticker, strategy_variant_id, direction, entry, exit_)
    return {
        "session_date": session_date,
        "evidence_status": evidence_status,
        "canonical_outcome_id": canonical_id,
        "ticker": ticker,
        "broad_sector": None if broad_sector is None or pd.isna(broad_sector) else str(broad_sector),
        "strategy_variant_id": strategy_variant_id,
        "direction": None if direction is None or pd.isna(direction) else str(direction).upper(),
        "morning_available_flag": bool(morning_available),
        "governance_feasible_flag": bool(governance_feasible),
        "outcome_status": outcome_status,
        "entry_time": entry,
        "exit_time": exit_,
        "native_notional_sek": _safe_float(native_notional),
        "native_net_pnl_sek": _safe_float(native_pnl),
        "direction_adjusted_gross_return": _safe_float(gross_return),
        "standardized_cost_sek": _safe_float(standardized_cost),
        "standardized_net_pnl_sek": _safe_float(standardized_pnl),
        "mfe_return": _safe_float(mfe),
        "mae_return": _safe_float(mae),
        "authoritative_source": source,
        "source_engine_memberships": "|".join(sorted(set(str(x) for x in memberships))),
    }


def _dim_engine() -> pd.DataFrame:
    rows = [
        ("STEP9I_V2_SELECTED", "STEP9I", "SELECTED", True, 2, None, "ACTUAL_SELECTED"),
        ("STEP9L_V3_SELECTED", "STEP9L", "SELECTED", True, 2, None, "ACTUAL_SELECTED"),
        ("STEP9S_NATURAL", "STEP9S", "NATURAL", True, 2, None, "NATURAL_BOOK"),
        ("STEP9S_MANDATORY_CONTROL", "STEP9S", "MANDATORY_CONTROL", True, 1, None, "MANDATORY_CONTROL"),
        ("STEP9R_SELECTED", "STEP9R", "SELECTED", True, 2, None, "ACTUAL_SELECTED"),
        ("STEP9U_SELECTED", "STEP9U", "SELECTED", True, 2, 1, "ACTUAL_SELECTED"),
        ("STEP9U_PLUS_STEP9V_ACTION", "STEP9V", "OBSERVER", False, 2, 1, "OBSERVER_COUNTERFACTUAL"),
        ("STEP9T_ALL_DIRECTIONAL", "STEP9T", "OBSERVER", False, None, None, "CANDIDATE_COUNTERFACTUAL"),
    ]
    return pd.DataFrame(rows, columns=[
        "engine_book_id", "engine_id", "book_type", "main_chart_eligible",
        "max_positions", "max_positions_per_sector", "result_type",
    ])


def _dim_strategy(project_root: Path, config: dict[str, Any], outcomes: pd.DataFrame) -> pd.DataFrame:
    rows: dict[str, dict[str, Any]] = {}

    registry_path = project_root / config["optional_registry_paths"]["step9l_contract_registry"]
    if registry_path.exists():
        registry = pd.read_csv(registry_path)
        for record in registry.to_dict("records"):
            variant = f"STEP9L::{record['contract_id']}"
            rows[variant] = {
                "strategy_variant_id": variant,
                "strategy_family": str(record.get("base_challenger_id") or record["contract_id"]),
                "source_engine_id": "STEP9L",
                "direction_model": "DYNAMIC",
                "entry_model": str(record.get("base_challenger_id") or "CONTRACT_DEFINED"),
                "exit_model": "CONTRACT_DEFINED_STOP_TARGET_TIME_EXIT",
                "applicable_regimes": str(record.get("primary_regime") or ""),
                "governance_status": "SELECTABLE" if str(record.get("test_role")) == "PRIMARY_HYPOTHESIS" else "BLOCKED",
                "comparison_eligible": str(record.get("test_role")) == "PRIMARY_HYPOTHESIS",
            }

    for variant, regimes in config.get("strategy_archetype_regime_registry", {}).items():
        archetype = variant.split("::")[1] if "::" in variant else variant
        direction = "LONG" if archetype.endswith("LONG") else "SHORT" if archetype.endswith("SHORT") else "DYNAMIC"
        rows[variant] = {
            "strategy_variant_id": variant,
            "strategy_family": archetype,
            "source_engine_id": "STEP9T",
            "direction_model": direction,
            "entry_model": "STANDARDIZED_0950_ENTRY",
            "exit_model": "EOD_CLOSE",
            "applicable_regimes": "|".join(regimes),
            "governance_status": "OBSERVATION_ONLY",
            "comparison_eligible": True,
        }

    if not outcomes.empty:
        for variant in outcomes["strategy_variant_id"].dropna().astype(str).unique():
            if variant in rows:
                continue
            if variant.startswith("STEP9S_CONTROL::"):
                rows[variant] = {
                    "strategy_variant_id": variant,
                    "strategy_family": variant.split("::", 1)[1],
                    "source_engine_id": "STEP9S",
                    "direction_model": "DYNAMIC",
                    "entry_model": "CONTROL_DEFINED_0950_WINDOW",
                    "exit_model": "CONTROL_DEFINED_STOP_TARGET_TIME_EXIT",
                    "applicable_regimes": "",
                    "governance_status": "CONTROL",
                    "comparison_eligible": True,
                }
            elif variant.startswith("STEP9V::"):
                rows[variant] = {
                    "strategy_variant_id": variant,
                    "strategy_family": "INTRADAY_MANAGEMENT",
                    "source_engine_id": "STEP9V",
                    "direction_model": "DYNAMIC",
                    "entry_model": "MORNING_ENTRY_PLUS_CHECKPOINT_ACTION",
                    "exit_model": variant.rsplit("::", 1)[-1],
                    "applicable_regimes": "",
                    "governance_status": "OBSERVATION_ONLY",
                    "comparison_eligible": False,
                }
            elif variant.startswith("STEP9L::"):
                contract = variant.split("::", 1)[1]
                rows[variant] = {
                    "strategy_variant_id": variant,
                    "strategy_family": contract,
                    "source_engine_id": "STEP9L",
                    "direction_model": "DYNAMIC",
                    "entry_model": "CONTRACT_DEFINED",
                    "exit_model": "CONTRACT_DEFINED",
                    "applicable_regimes": "",
                    "governance_status": "SELECTABLE",
                    "comparison_eligible": True,
                }

    rows["CASH_NO_TRADE"] = {
        "strategy_variant_id": "CASH_NO_TRADE",
        "strategy_family": "Cash / no trade",
        "source_engine_id": "CASH",
        "direction_model": "CASH",
        "entry_model": "NONE",
        "exit_model": "NONE",
        "applicable_regimes": "*",
        "governance_status": "CASH",
        "comparison_eligible": True,
    }
    return pd.DataFrame(sorted(rows.values(), key=lambda r: r["strategy_variant_id"]))


def _engine_daily(
    sources: dict[str, SourceSpec],
    evidence: dict[str, str],
    config: dict[str, Any],
) -> pd.DataFrame:
    notional = float(config["comparison_notional_sek"])
    cost_rate = float(config["comparison_round_trip_cost_rate"])
    rows: list[dict[str, Any]] = []

    def aggregate(
        engine_book: str,
        session: str,
        frame: pd.DataFrame,
        native_column: str,
        gross_column: str,
        selected_mask: pd.Series,
        outcome_batch_id: Any,
        selected_count_override: int | None = None,
    ) -> None:
        selected = frame.loc[selected_mask].copy()
        completed = selected.loc[selected.apply(lambda r: _normalize_outcome_status(r.get("outcome_status"), r.get(gross_column)) == "COMPLETE", axis=1)]
        no_trigger = selected.loc[selected.apply(lambda r: _normalize_outcome_status(r.get("outcome_status"), r.get(gross_column)) == "NO_TRIGGER", axis=1)]
        standardized_values = []
        for value in completed[gross_column].tolist() if gross_column in completed.columns else []:
            _, pnl = _standardize_return(value, notional, cost_rate)
            if pnl is not None:
                standardized_values.append(pnl)
        native_values = pd.to_numeric(completed[native_column], errors="coerce") if native_column in completed.columns else pd.Series(dtype=float)
        selected_count = selected_count_override if selected_count_override is not None else len(selected)
        pnl = float(np.nansum(standardized_values)) if standardized_values else 0.0
        deployed = len(completed) * notional
        winners = sum(1 for value in standardized_values if value > 0)
        losses = sum(1 for value in standardized_values if value < 0)
        rows.append({
            "session_date": session,
            "evidence_status": evidence.get(session, "PROSPECTIVE_EXCLUDED"),
            "engine_book_id": engine_book,
            "run_status": "COMPLETE" if selected_count else "VALID_NO_TRADE",
            "selected_count": int(selected_count),
            "completed_trade_count": int(len(completed)),
            "winner_count": int(winners),
            "loss_count": int(losses),
            "no_trigger_count": int(len(no_trigger)),
            "native_net_pnl_sek": float(native_values.sum(skipna=True)) if not native_values.empty else 0.0,
            "standardized_net_pnl_sek": pnl,
            "capital_deployed_sek": float(deployed),
            "return_on_deployed_capital": None if deployed == 0 else pnl / deployed,
            "source_outcome_batch_id": None if outcome_batch_id is None or pd.isna(outcome_batch_id) else str(outcome_batch_id),
        })

    for source_id, engine_book in (("step9i", "STEP9I_V2_SELECTED"), ("step9l", "STEP9L_V3_SELECTED")):
        path = sources[source_id].path
        if not path.exists():
            continue
        outcomes = _read_table(path, "shadow_outcomes")
        batches = _read_table(path, "shadow_outcome_batches")
        for session in sorted(outcomes["session_date"].astype(str).unique()) if not outcomes.empty else []:
            frame = outcomes.loc[outcomes["session_date"].astype(str).eq(session)]
            # Actual selected book excludes guardrails and complement controls.
            mask = frame["selected_for_simulation"].fillna(0).astype(int).eq(1) & frame["test_role"].astype(str).eq("PRIMARY_HYPOTHESIS")
            batch_id = None
            if not batches.empty:
                match = batches.loc[batches["session_date"].astype(str).eq(session)]
                batch_id = match.iloc[-1].get("outcome_batch_id") if not match.empty else None
            aggregate(engine_book, session, frame, "risk_capped_net_pnl_sek", "gross_return", mask, batch_id)

    if sources["step9s"].path.exists():
        batches = _read_table(sources["step9s"].path, "step9s_outcome_batches")
        natural = _read_table(sources["step9s"].path, "step9s_natural_outcomes")
        control = _read_table(sources["step9s"].path, "step9s_coverage_outcomes")
        sessions = sorted(set(natural.get("session_date", pd.Series(dtype=str)).astype(str)) | set(control.get("session_date", pd.Series(dtype=str)).astype(str)))
        for session in sessions:
            batch = batches.loc[batches["session_date"].astype(str).eq(session)] if not batches.empty else pd.DataFrame()
            batch_id = batch.iloc[-1].get("outcome_batch_id") if not batch.empty else None
            nat = natural.loc[natural["session_date"].astype(str).eq(session)] if not natural.empty else natural
            ctl = control.loc[control["session_date"].astype(str).eq(session)] if not control.empty else control
            aggregate("STEP9S_NATURAL", session, nat.assign(outcome_status="COMPLETE"), "net_pnl_sek", "gross_return", pd.Series(True, index=nat.index), batch_id)
            aggregate("STEP9S_MANDATORY_CONTROL", session, ctl.assign(outcome_status="COMPLETE"), "net_pnl_sek", "gross_return", pd.Series(True, index=ctl.index), batch_id)

    if sources["step9r"].path.exists():
        batches = _read_table(sources["step9r"].path, "selector_batches")
        candidates = _read_table(sources["step9r"].path, "selector_candidates")
        outcomes = _read_table(sources["step9r"].path, "selector_outcomes")
        for session in sorted(batches["session_date"].astype(str).unique()) if not batches.empty else []:
            batch = batches.loc[batches["session_date"].astype(str).eq(session)].iloc[-1]
            selected_count = int(batch.get("selected_rows", 0))
            frame = outcomes.loc[outcomes["session_date"].astype(str).eq(session)].copy() if not outcomes.empty else pd.DataFrame()
            if frame.empty:
                frame = pd.DataFrame(columns=["outcome_status", "risk_capped_net_pnl_sek", "net_r_after_costs", "selected"])
            frame["gross_return"] = pd.to_numeric(frame.get("risk_capped_net_pnl_sek"), errors="coerce") / 750.0 + cost_rate
            frame["outcome_status"] = np.where(frame["risk_capped_net_pnl_sek"].notna(), "COMPLETE", "INCOMPLETE")
            aggregate("STEP9R_SELECTED", session, frame, "risk_capped_net_pnl_sek", "gross_return", frame.get("selected", pd.Series(dtype=int)).fillna(0).astype(int).eq(1), batch.get("batch_id"), selected_count_override=selected_count)

    if sources["step9u"].path.exists():
        batches = _read_table(sources["step9u"].path, "step9u_prospective_outcome_batches")
        outcomes = _read_table(sources["step9u"].path, "step9u_prospective_candidate_outcomes")
        for session in sorted(outcomes["session_date"].astype(str).unique()) if not outcomes.empty else []:
            frame = outcomes.loc[outcomes["session_date"].astype(str).eq(session)].copy()
            batch = batches.loc[batches["session_date"].astype(str).eq(session)] if not batches.empty else pd.DataFrame()
            batch_id = batch.iloc[-1].get("outcome_batch_id") if not batch.empty else None
            aggregate("STEP9U_SELECTED", session, frame, "net_pnl_sek", "session_close_return", frame["selected"].fillna(0).astype(int).eq(1), batch_id)

    if sources["step9t"].path.exists():
        batches = _read_table(sources["step9t"].path, "step9t_prospective_outcome_batches")
        outcomes = _read_table(sources["step9t"].path, "step9t_prospective_ticker_outcomes")
        for session in sorted(outcomes["session_date"].astype(str).unique()) if not outcomes.empty else []:
            frame = outcomes.loc[outcomes["session_date"].astype(str).eq(session)].copy()
            frame = frame.loc[frame["direction"].astype(str).isin(["LONG", "SHORT"])]
            batch = batches.loc[batches["session_date"].astype(str).eq(session)] if not batches.empty else pd.DataFrame()
            batch_id = batch.iloc[-1].get("outcome_batch_id") if not batch.empty else None
            aggregate("STEP9T_ALL_DIRECTIONAL", session, frame, "net_pnl_sek", "session_close_return", pd.Series(True, index=frame.index), batch_id)

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values(["session_date", "engine_book_id"]).reset_index(drop=True)


def _best_per_ticker(outcomes: pd.DataFrame, feasible: bool, strategy_dim: pd.DataFrame) -> pd.DataFrame:
    if outcomes.empty:
        return outcomes.copy()
    dim = strategy_dim[["strategy_variant_id", "comparison_eligible"]]
    frame = outcomes.merge(dim, on="strategy_variant_id", how="left")
    frame = frame.loc[
        frame["outcome_status"].isin(["COMPLETE", "NO_TRIGGER"])
        & frame["standardized_net_pnl_sek"].notna()
        & frame["comparison_eligible"].fillna(False)
    ].copy()
    if feasible:
        frame = frame.loc[frame["morning_available_flag"] & frame["governance_feasible_flag"]]
    frame = frame.sort_values(["standardized_net_pnl_sek", "strategy_variant_id"], ascending=[False, True])
    return frame.drop_duplicates("ticker", keep="first")


def _select_portfolio(
    candidates: pd.DataFrame,
    n: int,
    fixed: bool,
    positive_only: bool,
    max_positions_per_sector: int | None = None,
    rank_column: str | None = None,
) -> tuple[pd.DataFrame, str]:
    if n == 0:
        return candidates.iloc[0:0].copy(), "COMPLETE"
    if candidates.empty:
        return candidates.copy(), "NOT_EVALUABLE"
    frame = candidates.copy()
    if positive_only:
        frame = frame.loc[frame["standardized_net_pnl_sek"] > 0]
    if rank_column:
        frame = frame.loc[frame[rank_column].notna()].sort_values([rank_column, "ticker"], ascending=[True, True])
    else:
        frame = frame.sort_values(["standardized_net_pnl_sek", "ticker"], ascending=[False, True])

    selected_rows: list[pd.Series] = []
    used_tickers: set[str] = set()
    sector_counts: dict[str, int] = {}
    for _, row in frame.iterrows():
        ticker = str(row["ticker"])
        if ticker in used_tickers:
            continue
        sector = str(row.get("broad_sector") or "UNKNOWN")
        if max_positions_per_sector is not None and sector_counts.get(sector, 0) >= max_positions_per_sector:
            continue
        selected_rows.append(row)
        used_tickers.add(ticker)
        sector_counts[sector] = sector_counts.get(sector, 0) + 1
        if len(selected_rows) >= n:
            break
    result = pd.DataFrame(selected_rows, columns=frame.columns) if selected_rows else frame.iloc[0:0].copy()
    if fixed and len(result) < n:
        return result, "NOT_EVALUABLE"
    return result, "COMPLETE"


def _benchmarks(outcomes: pd.DataFrame, strategy_dim: pd.DataFrame, evidence: dict[str, str]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for session in sorted(outcomes["session_date"].astype(str).unique()) if not outcomes.empty else []:
        session_outcomes = outcomes.loc[outcomes["session_date"].astype(str).eq(session)]
        # The approved specification contains both an unrestricted algorithm in section 5
        # and a sector-capped July 29 fixture in section 13. Emit both interpretations
        # rather than silently discarding either one. The requested main benchmark is
        # the truly unrestricted ORACLE_TOP2_OBSERVED_FIXED series.
        variants = [
            ("ORACLE_TOP2_OBSERVED_FIXED", False, True, False, None),
            ("ORACLE_UP_TO2_OBSERVED_NO_FORCE", False, False, True, None),
            ("ORACLE_TOP2_OBSERVED_SECTOR_CAPPED_FIXED", False, True, False, 1),
            ("ORACLE_TOP2_FEASIBLE_FIXED", True, True, False, 1),
            ("ORACLE_UP_TO2_FEASIBLE_NO_FORCE", True, False, True, 1),
        ]
        for benchmark_id, feasible, fixed, positive_only, sector_cap in variants:
            best = _best_per_ticker(session_outcomes, feasible, strategy_dim)
            selected, status = _select_portfolio(best, 2, fixed, positive_only, sector_cap)
            rows.append({
                "session_date": session,
                "evidence_status": evidence.get(session, "PROSPECTIVE_EXCLUDED"),
                "benchmark_id": benchmark_id,
                "fixed_or_up_to": "FIXED" if fixed else "UP_TO",
                "feasibility_scope": "MORNING_FEASIBLE" if feasible else "OBSERVED_UNRESTRICTED",
                "requested_position_count": 2,
                "actual_position_count": len(selected),
                "selected_tickers": "|".join(selected["ticker"].astype(str)) if not selected.empty else "",
                "selected_strategy_variants": "|".join(selected["strategy_variant_id"].astype(str)) if not selected.empty else "",
                "standardized_net_pnl_sek": None if status != "COMPLETE" else float(selected["standardized_net_pnl_sek"].sum()),
                "candidate_ticker_count": int(best["ticker"].nunique()) if not best.empty else 0,
                "strategy_outcome_count": int(len(session_outcomes.loc[session_outcomes["standardized_net_pnl_sek"].notna()])),
                "coverage_status": "COMPLETE" if status == "COMPLETE" else "NOT_EVALUABLE",
            })
    return pd.DataFrame(rows)


def _canonical_applied_strategy_map(
    engine_book: str,
    session: str,
    sources: dict[str, SourceSpec],
) -> tuple[set[str], dict[str, str], dict[str, bool]]:
    universe: set[str] = set()
    applied: dict[str, str] = {}
    selected: dict[str, bool] = {}
    if engine_book in {"STEP9I_V2_SELECTED", "STEP9L_V3_SELECTED"}:
        source = "step9i" if engine_book.startswith("STEP9I") else "step9l"
        decisions = _read_table(sources[source].path, "shadow_decisions", "session_date=?", (session,))
        outcomes = _read_table(sources[source].path, "shadow_outcomes", "session_date=?", (session,))
        if not decisions.empty:
            universe = set(decisions["ticker"].astype(str).unique())
        if not outcomes.empty:
            sel = outcomes.loc[
                outcomes["selected_for_simulation"].fillna(0).astype(int).eq(1)
                & outcomes["test_role"].astype(str).eq("PRIMARY_HYPOTHESIS")
            ]
            for row in sel.to_dict("records"):
                ticker = str(row["ticker"])
                applied[ticker] = f"STEP9L::{row['contract_id']}"
                selected[ticker] = True
    elif engine_book == "STEP9R_SELECTED" and sources["step9r"].path.exists():
        candidates = _read_table(sources["step9r"].path, "selector_candidates", "session_date=?", (session,))
        universe = set(candidates["ticker"].astype(str).unique()) if not candidates.empty else set()
        for row in candidates.loc[candidates.get("selected", 0).fillna(0).astype(int).eq(1)].to_dict("records") if not candidates.empty else []:
            ticker = str(row["ticker"])
            applied[ticker] = f"STEP9L::{row['contract_id']}"
            selected[ticker] = True
    elif engine_book == "STEP9U_SELECTED" and sources["step9u"].path.exists():
        candidates = _read_table(sources["step9u"].path, "step9u_prospective_candidates", "session_date=?", (session,))
        universe = set(candidates["ticker"].astype(str).unique()) if not candidates.empty else set()
        for row in candidates.loc[candidates.get("selected", 0).fillna(0).astype(int).eq(1)].to_dict("records") if not candidates.empty else []:
            ticker = str(row["ticker"])
            applied[ticker] = f"STEP9T::{row['primary_archetype']}::0950_TO_EOD"
            selected[ticker] = True
    return universe, applied, selected


def _strategy_accuracy(
    outcomes: pd.DataFrame,
    strategy_dim: pd.DataFrame,
    sources: dict[str, SourceSpec],
    evidence: dict[str, str],
    tie_tolerance: float,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    comparison_eligible_ids = set(
        strategy_dim.loc[strategy_dim["comparison_eligible"].fillna(False).astype(bool), "strategy_variant_id"].astype(str)
    ) if not strategy_dim.empty else set()
    engine_books = ["STEP9I_V2_SELECTED", "STEP9L_V3_SELECTED", "STEP9R_SELECTED", "STEP9U_SELECTED"]
    for session in sorted(evidence):
        alternatives = outcomes.loc[
            outcomes["session_date"].astype(str).eq(session)
            & outcomes["outcome_status"].isin(["COMPLETE", "NO_TRIGGER"])
            & outcomes["standardized_net_pnl_sek"].notna()
            & outcomes["strategy_variant_id"].astype(str).isin(comparison_eligible_ids)
        ]
        for engine_book in engine_books:
            universe, applied_map, selected_map = _canonical_applied_strategy_map(engine_book, session, sources)
            if not universe:
                continue
            for ticker in sorted(universe):
                ticker_outcomes = alternatives.loc[alternatives["ticker"].astype(str).eq(ticker)].copy()
                cash_row = {"strategy_variant_id": "CASH_NO_TRADE", "standardized_net_pnl_sek": 0.0}
                options = ticker_outcomes[["strategy_variant_id", "standardized_net_pnl_sek"]].to_dict("records") + [cash_row]
                best_pnl = max(float(item["standardized_net_pnl_sek"]) for item in options)
                tie_ids = sorted(str(item["strategy_variant_id"]) for item in options if best_pnl - float(item["standardized_net_pnl_sek"]) <= tie_tolerance)
                best_id = tie_ids[0]
                applied_id = applied_map.get(ticker, "CASH_NO_TRADE")
                applied_rows = ticker_outcomes.loc[ticker_outcomes["strategy_variant_id"].eq(applied_id)]
                applied_pnl = 0.0 if applied_id == "CASH_NO_TRADE" else (_safe_float(applied_rows.iloc[0]["standardized_net_pnl_sek"]) if not applied_rows.empty else None)
                if applied_pnl is None:
                    state = "NOT_EVALUABLE"
                    correct: bool | None = None
                    opportunity = None
                else:
                    correct = applied_id in tie_ids
                    opportunity = max(0.0, best_pnl - applied_pnl)
                    if applied_id == "CASH_NO_TRADE":
                        state = "NO_TRADE_CORRECT" if correct else "NO_TRADE_FALSE_NEGATIVE"
                    elif best_id == "CASH_NO_TRADE" and applied_pnl < 0:
                        state = "TRADE_FALSE_POSITIVE_VS_CASH"
                    elif correct and len(tie_ids) > 1:
                        state = "APPLIED_TIED_BEST_STRATEGY"
                    elif correct:
                        state = "APPLIED_BEST_STRATEGY"
                    else:
                        state = "APPLIED_NOT_BEST"
                rows.append({
                    "session_date": session,
                    "evidence_status": evidence[session],
                    "engine_book_id": engine_book,
                    "ticker": ticker,
                    "decision_universe_flag": True,
                    "selected_flag": bool(selected_map.get(ticker, False)),
                    "applied_strategy_variant_id": applied_id,
                    "applied_standardized_pnl_sek": applied_pnl,
                    "oracle_best_strategy_variant_id": best_id,
                    "oracle_best_standardized_pnl_sek": best_pnl,
                    "strategy_accuracy_state": state,
                    "strategy_correct_flag": correct,
                    "strategy_opportunity_loss_sek": opportunity,
                    "evaluated_strategy_count": len(options),
                })
    return pd.DataFrame(rows)


def _canonical_prices(price_db: Path) -> pd.DataFrame:
    with _connect_ro(price_db) as connection:
        query = """
        SELECT substr(datetime,1,16) || ":00" AS datetime, open, high, low, close, ticker, source, collected_at_utc
        FROM (
          SELECT *, ROW_NUMBER() OVER (
            PARTITION BY ticker, substr(datetime,1,16)
            ORDER BY rowid DESC
          ) AS rn
          FROM intraday_prices
        )
        WHERE rn=1
        ORDER BY datetime, ticker
        """
        frame = pd.read_sql_query(query, connection)
    frame["datetime"] = pd.to_datetime(frame["datetime"], errors="coerce", format="mixed")
    frame = frame.dropna(subset=["datetime", "ticker", "open", "high", "low", "close"])
    frame["session_date"] = frame["datetime"].dt.strftime("%Y-%m-%d")
    frame["time_label"] = frame["datetime"].dt.strftime("%H:%M")
    return frame


def _percentile_against_prior(current: float | None, prior: list[float], min_prior: int) -> float | None:
    if current is None or len(prior) < min_prior:
        return None
    array = np.asarray([x for x in prior if x is not None and math.isfinite(x)], dtype=float)
    if len(array) < min_prior:
        return None
    return float(np.mean(array <= current))


def _session_regime_features(prices: pd.DataFrame, tickers: list[str], session: str, config: dict[str, Any]) -> dict[str, Any]:
    rcfg = config["realized_regime"]
    opening_label = rcfg["opening_label"]
    morning_label = rcfg["morning_label"]
    eod_min = rcfg["eod_minimum_label"]
    day = prices.loc[prices["session_date"].eq(session) & prices["ticker"].isin(tickers)].copy()
    if day.empty:
        return {"session_date": session, "required_ticker_count": len(tickers), "valid_ticker_count": 0}
    eod_labels = sorted(label for label in day["time_label"].unique() if label >= eod_min)
    eod_label = eod_labels[-1] if eod_labels else None
    previous_dates = sorted(d for d in prices["session_date"].unique() if d < session)
    previous = previous_dates[-1] if previous_dates else None
    prev_day = prices.loc[prices["session_date"].eq(previous) & prices["ticker"].isin(tickers)] if previous else pd.DataFrame()

    records = []
    for ticker in tickers:
        td = day.loc[day["ticker"].eq(ticker)].sort_values("datetime")
        if td.empty:
            continue
        open_rows = td.loc[td["time_label"].eq(opening_label)]
        morning_rows = td.loc[td["time_label"].eq(morning_label)]
        eod_rows = td.loc[td["time_label"].eq(eod_label)] if eod_label else pd.DataFrame()
        if open_rows.empty or morning_rows.empty or eod_rows.empty:
            continue
        open_price = float(open_rows.iloc[-1]["open"])
        morning_price = float(morning_rows.iloc[-1]["close"])
        eod_price = float(eod_rows.iloc[-1]["close"])
        prev_close = None
        if previous and not prev_day.empty:
            pt = prev_day.loc[prev_day["ticker"].eq(ticker)].sort_values("datetime")
            if not pt.empty:
                prev_close = float(pt.iloc[-1]["close"])
        returns = td.loc[td["time_label"].le(eod_label), "close"].astype(float).pct_change().dropna() if eod_label else pd.Series(dtype=float)
        records.append({
            "ticker": ticker,
            "opening_gap": None if not prev_close else open_price / prev_close - 1.0,
            "morning_return": morning_price / open_price - 1.0,
            "eod_return": eod_price / open_price - 1.0,
            "post_morning_return": eod_price / morning_price - 1.0,
            "range_return": (float(td["high"].max()) - float(td["low"].min())) / open_price,
            "realized_volatility": float(np.sqrt(np.square(returns).sum())) if len(returns) else 0.0,
        })
    ticker_df = pd.DataFrame(records)
    required = len(tickers)
    valid = len(ticker_df)
    coverage = valid / required if required else 0.0
    if ticker_df.empty:
        return {"session_date": session, "required_ticker_count": required, "valid_ticker_count": 0, "coverage": coverage, "eod_label": eod_label}

    median_eod = float(ticker_df["eod_return"].median())
    market_returns = []
    if eod_label:
        pivot = day.loc[day["time_label"].le(eod_label)].pivot_table(index="datetime", columns="ticker", values="close", aggfunc="last")
        ew = pivot.pct_change(fill_method=None).mean(axis=1, skipna=True).dropna()
        direction = np.sign(median_eod)
        nonzero = ew.loc[np.sign(ew) != 0]
        persistence = None if direction == 0 or nonzero.empty else float(np.mean(np.sign(nonzero) == direction))
    else:
        persistence = None

    morning_median = float(ticker_df["morning_return"].median())
    if abs(morning_median) >= 0.001 and abs(median_eod) >= 0.001 and np.sign(morning_median) != np.sign(median_eod):
        reversal = 1.0
    elif np.sign(morning_median) == np.sign(median_eod) and abs(morning_median) > 0 and abs(median_eod) < abs(morning_median):
        reversal = 1.0 - abs(median_eod) / abs(morning_median)
    else:
        reversal = 0.0

    return {
        "session_date": session,
        "required_ticker_count": required,
        "valid_ticker_count": valid,
        "coverage": coverage,
        "eod_label": eod_label,
        "median_opening_gap": (
            _safe_float(pd.to_numeric(ticker_df["opening_gap"], errors="coerce").dropna().median())
            if not pd.to_numeric(ticker_df["opening_gap"], errors="coerce").dropna().empty
            else None
        ),
        "gap_down_share": float(np.mean(pd.to_numeric(ticker_df["opening_gap"], errors="coerce").fillna(0.0) < 0)),
        "gap_up_share": float(np.mean(pd.to_numeric(ticker_df["opening_gap"], errors="coerce").fillna(0.0) > 0)),
        "median_morning_return": morning_median,
        "median_eod_return": median_eod,
        "median_post_morning_return": float(ticker_df["post_morning_return"].median()),
        "advancer_share": float(np.mean(ticker_df["eod_return"] > 0)),
        "decliner_share": float(np.mean(ticker_df["eod_return"] < 0)),
        "median_range": float(ticker_df["range_return"].median()),
        "median_realized_volatility": float(ticker_df["realized_volatility"].median()),
        "dispersion": float(ticker_df["eod_return"].std(ddof=0)),
        "path_persistence": persistence,
        "reversal_strength": reversal,
    }


def _classify_regime(feature: dict[str, Any], prior_features: list[dict[str, Any]], config: dict[str, Any]) -> dict[str, Any]:
    rcfg = config["realized_regime"]
    min_prior = int(rcfg["minimum_prior_sessions"])
    range_pct = _percentile_against_prior(feature.get("median_range"), [x.get("median_range") for x in prior_features], min_prior)
    vol_pct = _percentile_against_prior(feature.get("median_realized_volatility"), [x.get("median_realized_volatility") for x in prior_features], min_prior)
    disp_pct = _percentile_against_prior(feature.get("dispersion"), [x.get("dispersion") for x in prior_features], min_prior)
    max_vol_pct = max(x for x in (range_pct, vol_pct) if x is not None) if any(x is not None for x in (range_pct, vol_pct)) else None
    coverage = feature.get("coverage", 0.0)
    eod = feature.get("median_eod_return")
    morning = feature.get("median_morning_return")
    adv = feature.get("advancer_share")
    dec = feature.get("decliner_share")
    persist = feature.get("path_persistence")
    reversal = feature.get("reversal_strength")
    gap = feature.get("median_opening_gap")
    gap_down = feature.get("gap_down_share")
    post = feature.get("median_post_morning_return")

    flags = {
        "DATA_LIMITED_DEFENSIVE": coverage < float(rcfg["minimum_coverage"]) or feature.get("eod_label") is None,
        "RECOVERY": _all([_le(gap, -0.003), _ge(gap_down, 0.55), _ge(eod, 0.002), _gt(post, 0), _ge(adv, 0.55)]),
        "HIGH_VOL_REVERSAL": _all([_ge(max_vol_pct, 0.75), _ge(reversal, 0.70), _ge(abs(morning) if morning is not None else None, 0.0015)]),
        "VOLATILITY_EXPANSION": _all([_ge(max_vol_pct, 0.75), _ge(abs(eod) if eod is not None else None, 0.004), _ge(max(adv or 0, dec or 0), 0.60)]),
        "TREND_UP": _all([_ge(eod, 0.004), _ge(adv, 0.60), _ge(persist, 0.60)]),
        "TREND_DOWN": _all([_le(eod, -0.004), _ge(dec, 0.60), _ge(persist, 0.60)]),
        "HIGH_DISPERSION": _all([_ge(disp_pct, 0.75), _between(adv, 0.35, 0.65)]),
        "RANGE_LOW_VOL": _all([_le(max_vol_pct, 0.25), _le(abs(eod) if eod is not None else None, 0.002), _between(adv, 0.35, 0.65), _le(disp_pct, 0.50)]),
        "DEFENSIVE_MIXED": True,
    }
    priority = [
        "DATA_LIMITED_DEFENSIVE", "RECOVERY", "HIGH_VOL_REVERSAL", "VOLATILITY_EXPANSION",
        "TREND_UP", "TREND_DOWN", "HIGH_DISPERSION", "RANGE_LOW_VOL", "DEFENSIVE_MIXED",
    ]
    winning = next(name for name in priority if flags[name])
    true_specialist = sum(bool(flags[name]) for name in priority[:-1])
    return {
        **feature,
        "range_percentile": range_pct,
        "volatility_percentile": vol_pct,
        "dispersion_percentile": disp_pct,
        "rule_flags": flags,
        "winning_rule_id": winning,
        "true_rule_count": true_specialist,
        "realized_eod_regime": winning,
    }


def _all(values: Iterable[bool]) -> bool:
    return all(values)


def _ge(value: Any, threshold: float) -> bool:
    number = _safe_float(value)
    return False if number is None else number >= threshold


def _gt(value: Any, threshold: float) -> bool:
    number = _safe_float(value)
    return False if number is None else number > threshold


def _le(value: Any, threshold: float) -> bool:
    number = _safe_float(value)
    return False if number is None else number <= threshold


def _between(value: Any, low: float, high: float) -> bool:
    number = _safe_float(value)
    return False if number is None else low <= number <= high


def _regime_accuracy(
    prices: pd.DataFrame,
    regime_tickers: list[str],
    broad_tickers: list[str],
    sessions: list[str],
    morning_regime: dict[str, str | None],
    evidence: dict[str, str],
    config: dict[str, Any],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    all_sessions = sorted(prices["session_date"].unique())
    lookback = int(config["realized_regime"]["percentile_lookback_sessions"])
    feature_cache: dict[tuple[str, str], dict[str, Any]] = {}
    for universe_name, tickers in (("REGIME_SOURCE", regime_tickers), ("BROAD29_DIAGNOSTIC", broad_tickers)):
        for session in all_sessions:
            feature_cache[(universe_name, session)] = _session_regime_features(prices, tickers, session, config)
        for session in sessions:
            current = feature_cache.get((universe_name, session), {})
            prior_dates = [d for d in all_sessions if d < session][-lookback:]
            prior = [feature_cache[(universe_name, d)] for d in prior_dates]
            result = _classify_regime(current, prior, config)
            morning = morning_regime.get(session)
            realized = result.get("realized_eod_regime")
            if morning is None:
                accuracy_state = "NOT_EVALUABLE"
            elif realized == "DATA_LIMITED_DEFENSIVE":
                accuracy_state = "EOD_DATA_LIMITED"
            elif morning == "DATA_LIMITED_DEFENSIVE":
                accuracy_state = "MORNING_DATA_LIMITED"
            else:
                accuracy_state = "EXACT_MATCH" if morning == realized else "MISMATCH"
            rule_payload = {
                "classifier_id": config["realized_regime"]["classifier_id"],
                "eod_label": result.get("eod_label"),
                "median_opening_gap": result.get("median_opening_gap"),
                "gap_down_share": result.get("gap_down_share"),
                "median_morning_return": result.get("median_morning_return"),
                "median_post_morning_return": result.get("median_post_morning_return"),
                "median_range": result.get("median_range"),
                "median_realized_volatility": result.get("median_realized_volatility"),
                "dispersion": result.get("dispersion"),
                "flags": result.get("rule_flags", {}),
            }
            rows.append({
                "session_date": session,
                "evidence_status": evidence.get(session, "PROSPECTIVE_EXCLUDED"),
                "classifier_universe": universe_name,
                "morning_regime": morning,
                "realized_eod_regime": realized,
                "regime_accuracy_state": accuracy_state,
                "winning_rule_id": result.get("winning_rule_id"),
                "true_rule_count": int(result.get("true_rule_count", 0)),
                "valid_ticker_count": int(result.get("valid_ticker_count", 0)),
                "required_ticker_count": int(result.get("required_ticker_count", len(tickers))),
                "median_eod_return": result.get("median_eod_return"),
                "advancer_share": result.get("advancer_share"),
                "decliner_share": result.get("decliner_share"),
                "range_percentile": result.get("range_percentile"),
                "volatility_percentile": result.get("volatility_percentile"),
                "dispersion_percentile": result.get("dispersion_percentile"),
                "path_persistence": result.get("path_persistence"),
                "reversal_strength": result.get("reversal_strength"),
                "rule_flags_json": json.dumps(rule_payload, sort_keys=True, separators=(",", ":")),
            })
    return pd.DataFrame(rows)


def _ranking_inputs(engine_book: str, session: str, sources: dict[str, SourceSpec], config: dict[str, Any]) -> pd.DataFrame:
    if engine_book == "STEP9R_SELECTED" and sources["step9r"].path.exists():
        candidates = _read_table(sources["step9r"].path, "selector_candidates", "session_date=?", (session,))
        outcomes = _read_table(sources["step9r"].path, "selector_candidate_outcomes", "session_date=?", (session,))
        if candidates.empty:
            return pd.DataFrame()
        merged = candidates.merge(outcomes[["candidate_id", "risk_capped_net_pnl_sek", "counterfactual_trade_generated", "exit_reason"]] if not outcomes.empty else pd.DataFrame(columns=["candidate_id", "risk_capped_net_pnl_sek", "counterfactual_trade_generated", "exit_reason"]), on="candidate_id", how="left")
        merged["predicted_score"] = pd.to_numeric(merged["simple_expected_r"], errors="coerce")
        merged["predicted_rank"] = pd.to_numeric(merged["research_rank"], errors="coerce")
        gross = pd.to_numeric(merged["risk_capped_net_pnl_sek"], errors="coerce") / 750.0 + float(config["comparison_round_trip_cost_rate"])
        merged["standardized_net_pnl_sek"] = gross * float(config["comparison_notional_sek"]) - float(config["comparison_notional_sek"]) * float(config["comparison_round_trip_cost_rate"])
        merged["outcome_status"] = np.where(merged["risk_capped_net_pnl_sek"].notna(), "COMPLETE", "INCOMPLETE")
        merged["broad_sector"] = None
        return merged[["ticker", "broad_sector", "predicted_score", "predicted_rank", "selected", "standardized_net_pnl_sek", "outcome_status"]]

    if engine_book == "STEP9U_SELECTED" and sources["step9u"].path.exists():
        candidates = _read_table(sources["step9u"].path, "step9u_prospective_candidates", "session_date=?", (session,))
        outcomes = _read_table(sources["step9u"].path, "step9u_prospective_candidate_outcomes", "session_date=?", (session,))
        if candidates.empty:
            return pd.DataFrame()
        merged = candidates.merge(outcomes[["candidate_id", "net_pnl_sek", "outcome_status"]] if not outcomes.empty else pd.DataFrame(columns=["candidate_id", "net_pnl_sek", "outcome_status"]), on="candidate_id", how="left", suffixes=("", "_outcome"))
        merged = merged.loc[merged["policy_action"].astype(str).eq("SELECTABLE_CHALLENGER")].copy()
        merged = merged.sort_values(["rule_priority", "signal_strength", "ticker"], ascending=[False, False, True]).reset_index(drop=True)
        merged["predicted_rank"] = np.arange(1, len(merged) + 1)
        merged["predicted_score"] = pd.to_numeric(merged["signal_strength"], errors="coerce")
        merged["standardized_net_pnl_sek"] = pd.to_numeric(merged["net_pnl_sek"], errors="coerce")
        merged["outcome_status"] = np.where(merged["standardized_net_pnl_sek"].notna(), "COMPLETE", "INCOMPLETE")
        return merged[["ticker", "broad_sector", "predicted_score", "predicted_rank", "selected", "standardized_net_pnl_sek", "outcome_status"]]
    return pd.DataFrame()


def _dense_and_ordinal_ranks(frame: pd.DataFrame, tolerance: float) -> pd.DataFrame:
    complete = frame.loc[frame["standardized_net_pnl_sek"].notna()].sort_values(["standardized_net_pnl_sek", "ticker"], ascending=[False, True]).copy()
    dense = 0
    previous = None
    dense_values = []
    for value in complete["standardized_net_pnl_sek"].astype(float):
        if previous is None or previous - value > tolerance:
            dense += 1
        dense_values.append(dense)
        previous = value
    complete["actual_dense_rank"] = dense_values
    complete["actual_ordinal_rank"] = np.arange(1, len(complete) + 1)
    return complete


def _ranking_tables(
    sources: dict[str, SourceSpec],
    evidence: dict[str, str],
    config: dict[str, Any],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    ticker_rows: list[dict[str, Any]] = []
    daily_rows: list[dict[str, Any]] = []
    portfolio_rows: list[dict[str, Any]] = []
    tolerance = float(config["pnl_tie_tolerance_sek"])
    for session in sorted(evidence):
        for engine_book, cap, sector_cap in (("STEP9R_SELECTED", 2, None), ("STEP9U_SELECTED", 2, 1)):
            frame = _ranking_inputs(engine_book, session, sources, config)
            if frame.empty:
                continue
            ranked = _dense_and_ordinal_ranks(frame, tolerance)
            merged = frame.merge(ranked[["ticker", "actual_dense_rank", "actual_ordinal_rank"]], on="ticker", how="left")
            merged["rank_error"] = merged["predicted_rank"] - merged["actual_ordinal_rank"]
            merged["absolute_rank_error"] = merged["rank_error"].abs()
            for row in merged.to_dict("records"):
                pnl = _safe_float(row.get("standardized_net_pnl_sek"))
                ticker_rows.append({
                    "session_date": session,
                    "evidence_status": evidence[session],
                    "engine_book_id": engine_book,
                    "ticker": str(row["ticker"]),
                    "broad_sector": row.get("broad_sector"),
                    "predicted_score": _safe_float(row.get("predicted_score")),
                    "predicted_rank": None if pd.isna(row.get("predicted_rank")) else int(row.get("predicted_rank")),
                    "actual_dense_rank": None if pd.isna(row.get("actual_dense_rank")) else int(row.get("actual_dense_rank")),
                    "actual_ordinal_rank": None if pd.isna(row.get("actual_ordinal_rank")) else int(row.get("actual_ordinal_rank")),
                    "rank_error": None if pd.isna(row.get("rank_error")) else int(row.get("rank_error")),
                    "absolute_rank_error": None if pd.isna(row.get("absolute_rank_error")) else int(row.get("absolute_rank_error")),
                    "selected_flag": _bool(row.get("selected")),
                    "standardized_net_pnl_sek": pnl,
                    "winner_flag": None if pnl is None else pnl > 0,
                    "outcome_status": str(row.get("outcome_status")),
                })

            complete = merged.loc[merged["standardized_net_pnl_sek"].notna()].copy()
            selected_actual = complete.loc[complete["selected"].fillna(0).astype(int).eq(1)]
            predicted_fixed, pred_status = _select_portfolio(complete, cap, True, False, sector_cap, "predicted_rank")
            oracle_fixed, oracle_fixed_status = _select_portfolio(complete, cap, True, False, sector_cap)
            oracle_up, oracle_up_status = _select_portfolio(complete, cap, False, True, sector_cap)
            predicted_top = set(predicted_fixed["ticker"].astype(str))
            actual_top = set(oracle_fixed["ticker"].astype(str))
            denominator = min(cap, len(complete))
            spearman = None
            if len(complete) >= 3:
                spearman = _safe_float(complete[["predicted_rank", "actual_ordinal_rank"]].corr(method="spearman").iloc[0, 1])
            predicted_top1 = complete.sort_values(["predicted_rank", "ticker"]).iloc[0]["ticker"] if not complete.empty else None
            realized_top1_tie = set(complete.loc[complete["actual_dense_rank"].eq(1), "ticker"].astype(str))
            actual_selected_pnl = float(selected_actual["standardized_net_pnl_sek"].sum()) if not selected_actual.empty else 0.0
            pred_pnl = None if pred_status != "COMPLETE" else float(predicted_fixed["standardized_net_pnl_sek"].sum())
            fixed_pnl = None if oracle_fixed_status != "COMPLETE" else float(oracle_fixed["standardized_net_pnl_sek"].sum())
            up_pnl = None if oracle_up_status != "COMPLETE" else float(oracle_up["standardized_net_pnl_sek"].sum())
            daily_rows.append({
                "session_date": session,
                "evidence_status": evidence[session],
                "engine_book_id": engine_book,
                "complete_ranked_candidate_count": len(complete),
                "selected_count": len(selected_actual),
                "spearman_rank_correlation": spearman,
                "top1_hit_flag": None if predicted_top1 is None else str(predicted_top1) in realized_top1_tie,
                "top2_overlap_count": len(predicted_top & actual_top) if denominator else None,
                "top2_overlap_pct": None if denominator == 0 else len(predicted_top & actual_top) / denominator,
                "actual_selected_pnl_sek": actual_selected_pnl,
                "predicted_top_cap_pnl_sek": pred_pnl,
                "oracle_fixed_cap_pnl_sek": fixed_pnl,
                "oracle_up_to_cap_no_force_pnl_sek": up_pnl,
                "ranking_regret_fixed_cap_sek": None if pred_pnl is None or fixed_pnl is None else fixed_pnl - pred_pnl,
                "total_selection_opportunity_loss_sek": None if up_pnl is None else max(0.0, up_pnl - actual_selected_pnl),
                "threshold_effect_vs_forced_rank_sek": None if pred_pnl is None else actual_selected_pnl - pred_pnl,
                "coverage_status": "COMPLETE" if len(complete) == len(frame) else "PARTIAL",
            })

            for mode in ("ENGINE_RANKED_FIXED_N", "ENGINE_POLICY_UP_TO_N", "ORACLE_FIXED_N", "ORACLE_UP_TO_N_NO_FORCE"):
                for n in config["portfolio_n_values"]:
                    if n == 0:
                        selected_n, status = complete.iloc[0:0].copy(), "COMPLETE"
                    elif mode == "ENGINE_RANKED_FIXED_N":
                        selected_n, status = _select_portfolio(complete, n, True, False, sector_cap, "predicted_rank")
                    elif mode == "ENGINE_POLICY_UP_TO_N":
                        policy = complete.loc[complete["selected"].fillna(0).astype(int).eq(1)].copy()
                        selected_n, status = _select_portfolio(policy, n, False, False, sector_cap, "predicted_rank")
                    elif mode == "ORACLE_FIXED_N":
                        selected_n, status = _select_portfolio(complete, n, True, False, sector_cap)
                    else:
                        selected_n, status = _select_portfolio(complete, n, False, True, sector_cap)
                    _append_portfolio_row(portfolio_rows, session, evidence[session], engine_book, mode, n, selected_n, status, sector_cap)
    return pd.DataFrame(ticker_rows), pd.DataFrame(daily_rows), pd.DataFrame(portfolio_rows)


def _append_portfolio_row(rows: list[dict[str, Any]], session: str, evidence: str, engine: str, mode: str, n: int, selected: pd.DataFrame, status: str, sector_cap: int | None) -> None:
    pnl = float(selected["standardized_net_pnl_sek"].sum()) if status == "COMPLETE" and not selected.empty else (0.0 if status == "COMPLETE" else None)
    actual = len(selected) if status == "COMPLETE" else len(selected)
    deployed = actual * 1000.0
    values = selected["standardized_net_pnl_sek"].astype(float).tolist() if not selected.empty else []
    rows.append({
        "session_date": session,
        "evidence_status": evidence,
        "engine_book_id": engine,
        "simulation_mode": mode,
        "position_count_n": int(n),
        "actual_position_count": int(actual),
        "selected_tickers": "|".join(selected["ticker"].astype(str)) if not selected.empty else "",
        "standardized_net_pnl_sek": pnl,
        "average_pnl_per_ticker_sek": None if actual == 0 else pnl / actual if pnl is not None else None,
        "capital_deployed_sek": deployed,
        "return_on_deployed_capital": None if deployed == 0 or pnl is None else pnl / deployed,
        "winner_count": sum(1 for x in values if x > 0),
        "loss_count": sum(1 for x in values if x < 0),
        "win_rate": None if not values else sum(1 for x in values if x > 0) / len(values),
        "worst_position_pnl_sek": None if not values else min(values),
        "constraints_applied": f"UNIQUE_TICKER;MAX_PER_SECTOR={sector_cap}" if sector_cap else "UNIQUE_TICKER",
        "evaluation_status": "COMPLETE" if status == "COMPLETE" else ("INSUFFICIENT_RANKED_CANDIDATES" if mode.startswith("ENGINE_RANKED") else "NOT_EVALUABLE"),
    })


def _official_regime_strategy_map(project_root: Path, config: dict[str, Any]) -> dict[str, str]:
    path = project_root / config["optional_registry_paths"]["step9s_config"]
    payload = _read_json(path, {}) or {}
    mapping: dict[str, str] = {}
    for item in payload.get("assignments", []):
        strategy = str(item.get("natural_strategy_id"))
        if strategy.startswith(("L_", "L2_", "L3_")):
            mapping[str(item["regime"])] = f"STEP9L::{strategy}"
        else:
            mapping[str(item["regime"])] = strategy
    return mapping


def _regime_strategy_accuracy(
    engine_daily: pd.DataFrame,
    outcomes: pd.DataFrame,
    strategy_dim: pd.DataFrame,
    regime_accuracy: pd.DataFrame,
    sources: dict[str, SourceSpec],
    project_root: Path,
    config: dict[str, Any],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    official_map = _official_regime_strategy_map(project_root, config)
    primary_regimes = regime_accuracy.loc[regime_accuracy["classifier_universe"].eq("REGIME_SOURCE")]
    strategy_lookup = strategy_dim.set_index("strategy_variant_id").to_dict("index") if not strategy_dim.empty else {}
    evaluated_books = ["STEP9L_V3_SELECTED", "STEP9S_NATURAL", "STEP9U_SELECTED"]
    for reg in primary_regimes.to_dict("records"):
        session = str(reg["session_date"])
        realized = reg.get("realized_eod_regime")
        morning = reg.get("morning_regime")
        session_outcomes = outcomes.loc[outcomes["session_date"].astype(str).eq(session)]
        compatible = []
        for row in session_outcomes.to_dict("records"):
            meta = strategy_lookup.get(str(row["strategy_variant_id"]), {})
            regimes = str(meta.get("applicable_regimes") or "").split("|")
            if realized in regimes and row.get("standardized_net_pnl_sek") is not None:
                compatible.append(row)
        compatible_df = pd.DataFrame(compatible)
        strategy_portfolios: list[tuple[str, float]] = []
        if not compatible_df.empty:
            for strategy_id, group in compatible_df.groupby("strategy_variant_id"):
                best = group.sort_values(["standardized_net_pnl_sek", "ticker"], ascending=[False, True]).drop_duplicates("ticker")
                selected, status = _select_portfolio(best, 2, False, True, 1)
                if status == "COMPLETE":
                    strategy_portfolios.append((str(strategy_id), float(selected["standardized_net_pnl_sek"].sum())))
        strategy_portfolios.sort(key=lambda x: (-x[1], x[0]))
        best_id = strategy_portfolios[0][0] if strategy_portfolios else None
        best_pnl = strategy_portfolios[0][1] if strategy_portfolios else None
        official_id = official_map.get(str(realized))
        official_pnl = next((pnl for sid, pnl in strategy_portfolios if sid == official_id), None)
        for engine_book in evaluated_books:
            daily = engine_daily.loc[
                engine_daily["session_date"].astype(str).eq(session)
                & engine_daily["engine_book_id"].eq(engine_book)
            ]
            if daily.empty:
                continue
            selected_pnl = _safe_float(daily.iloc[0].get("standardized_net_pnl_sek"))
            selected_variants: list[str] = []
            if engine_book in {"STEP9L_V3_SELECTED", "STEP9U_SELECTED"}:
                _, applied_map, selected_map = _canonical_applied_strategy_map(engine_book, session, sources)
                selected_variants = sorted({applied_map[t] for t, flag in selected_map.items() if flag and t in applied_map})
            elif engine_book == "STEP9S_NATURAL":
                selected_variants = [official_map.get(str(morning))] if official_map.get(str(morning)) else []
            compatible_flag = None
            if selected_variants and realized:
                compatible_flag = all(realized in str(strategy_lookup.get(v, {}).get("applicable_regimes") or "").split("|") for v in selected_variants)
            if best_pnl is None or selected_pnl is None:
                state = "NOT_EVALUABLE"
                loss = None
                coverage = "NOT_EVALUABLE"
            else:
                loss = max(0.0, best_pnl - selected_pnl)
                if best_pnl <= 0 and selected_pnl == 0:
                    state = "NO_TRADE_BEST"
                elif compatible_flag and loss <= float(config["pnl_tie_tolerance_sek"]):
                    state = "COMPATIBLE_AND_BEST"
                elif compatible_flag:
                    state = "COMPATIBLE_NOT_BEST"
                elif selected_pnl > 0:
                    state = "INCOMPATIBLE_BUT_PROFITABLE"
                else:
                    state = "INCOMPATIBLE_AND_NOT_BEST"
                coverage = "COMPLETE"
            rows.append({
                "session_date": session,
                "evidence_status": reg["evidence_status"],
                "engine_book_id": engine_book,
                "morning_regime": morning,
                "realized_eod_regime": realized,
                "selected_strategy_set": "|".join(v for v in selected_variants if v),
                "official_realized_regime_strategy_id": official_id,
                "best_compatible_strategy_id": best_id,
                "regime_compatible_flag": compatible_flag,
                "selected_strategy_pnl_sek": selected_pnl,
                "official_mapped_strategy_pnl_sek": official_pnl,
                "best_compatible_strategy_pnl_sek": best_pnl,
                "regime_strategy_opportunity_loss_sek": loss,
                "accuracy_state": state,
                "coverage_status": coverage,
            })
    return pd.DataFrame(rows)


def _dim_session(
    sessions: list[str],
    evidence: dict[str, str],
    morning_regime: dict[str, str | None],
    regime_accuracy: pd.DataFrame,
    prices: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    for session in sessions:
        primary = regime_accuracy.loc[
            regime_accuracy["session_date"].astype(str).eq(session)
            & regime_accuracy["classifier_universe"].eq("REGIME_SOURCE")
        ]
        broad = regime_accuracy.loc[
            regime_accuracy["session_date"].astype(str).eq(session)
            & regime_accuracy["classifier_universe"].eq("BROAD29_DIAGNOSTIC")
        ]
        day = prices.loc[prices["session_date"].eq(session)]
        last_bar = day["datetime"].max() if not day.empty else None
        valid = primary.iloc[0]["valid_ticker_count"] if not primary.empty else 0
        required = primary.iloc[0]["required_ticker_count"] if not primary.empty else 0
        rows.append({
            "session_date": session,
            "evidence_status": evidence.get(session, "PROSPECTIVE_EXCLUDED"),
            "session_status": "SEALED_COMPLETE",
            "morning_regime": morning_regime.get(session),
            "realized_eod_regime": None if primary.empty else primary.iloc[0]["realized_eod_regime"],
            "realized_eod_regime_broad29": None if broad.empty else broad.iloc[0]["realized_eod_regime"],
            "last_completed_bar": None if last_bar is None or pd.isna(last_bar) else pd.Timestamp(last_bar).isoformat(sep=" "),
            "source_coverage_pct": None if not required else float(valid) / float(required),
        })
    return pd.DataFrame(rows)


def _data_quality(
    sessions: list[str],
    evidence: dict[str, str],
    sources: dict[str, SourceSpec],
    source_hashes_before: dict[str, str],
    source_hashes_after: dict[str, str],
    tables: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for session in sessions:
        status = evidence.get(session, "PROSPECTIVE_EXCLUDED")
        for source_id, spec in sources.items():
            exists = spec.path.exists()
            same = source_hashes_before.get(source_id) == source_hashes_after.get(source_id)
            rows.append({
                "session_date": session,
                "evidence_status": status,
                "check_id": "SOURCE_READ_ONLY_HASH_UNCHANGED",
                "source_id": source_id,
                "passed": bool(exists and same),
                "rows_checked": None,
                "failure_count": 0 if exists and same else 1,
                "detail": "Source hash unchanged" if exists and same else "Missing source or source hash changed",
                "source_sha256_before": source_hashes_before.get(source_id),
                "source_sha256_after": source_hashes_after.get(source_id),
            })
        for name, frame in tables.items():
            if name == "tblDataQuality":
                continue
            session_rows = frame.loc[frame["session_date"].astype(str).eq(session)] if not frame.empty and "session_date" in frame.columns else frame
            rows.append({
                "session_date": session,
                "evidence_status": status,
                "check_id": "OUTPUT_TABLE_GENERATED",
                "source_id": name,
                "passed": True,
                "rows_checked": len(session_rows) if session_rows is not None else 0,
                "failure_count": 0,
                "detail": f"Generated {len(session_rows)} row(s) for session",
                "source_sha256_before": None,
                "source_sha256_after": None,
            })
        for source_id in ("step9s", "step9t", "step9u", "step9v"):
            spec = sources[source_id]
            if not spec.path.exists():
                continue
            violations = 0
            checked = 0
            with _connect_ro(spec.path) as connection:
                for table in _table_names(connection):
                    cols = {row[1] for row in connection.execute(f'PRAGMA table_info("{table}")')}
                    checks = []
                    if "router_active" in cols:
                        checks.append("COALESCE(router_active,0)<>0")
                    if "order_sent" in cols:
                        checks.append("COALESCE(order_sent,0)<>0")
                    if not checks:
                        continue
                    checked += connection.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
                    violations += connection.execute(f'SELECT COUNT(*) FROM "{table}" WHERE ' + " OR ".join(checks)).fetchone()[0]
            rows.append({
                "session_date": session,
                "evidence_status": status,
                "check_id": "ROUTER_AND_ORDERS_INACTIVE",
                "source_id": source_id,
                "passed": violations == 0,
                "rows_checked": checked,
                "failure_count": violations,
                "detail": "All router/order flags are zero" if violations == 0 else f"Found {violations} active flags",
                "source_sha256_before": source_hashes_before.get(source_id),
                "source_sha256_after": source_hashes_after.get(source_id),
            })
    return pd.DataFrame(rows)


def _validate_schema(tables: dict[str, pd.DataFrame], schema: dict[str, Any]) -> None:
    for table_name, contract in schema["tables"].items():
        if table_name not in tables:
            raise OutputContractError(f"Missing output table: {table_name}")
        frame = tables[table_name]
        expected = [c["name"] for c in contract["columns"]]
        missing = [name for name in expected if name not in frame.columns]
        extra = [name for name in frame.columns if name not in expected]
        if missing or extra:
            raise OutputContractError(f"{table_name} schema mismatch. Missing={missing}; extra={extra}")
        if contract.get("primary_key") and not frame.empty:
            key = contract["primary_key"]
            if frame.duplicated(key).any():
                duplicates = frame.loc[frame.duplicated(key, keep=False), key].head(10).to_dict("records")
                raise OutputContractError(f"{table_name} primary-key duplicates: {duplicates}")
        for column in contract["columns"]:
            if not column["nullable"] and not frame.empty and frame[column["name"]].isna().any():
                raise OutputContractError(f"{table_name}.{column['name']} contains nulls but is non-nullable.")


def _conform_columns(frame: pd.DataFrame, schema: dict[str, Any], table_name: str) -> pd.DataFrame:
    columns = [c["name"] for c in schema["tables"][table_name]["columns"]]
    if frame.empty:
        return pd.DataFrame(columns=columns)
    for column in columns:
        if column not in frame.columns:
            frame[column] = None
    return frame[columns]


def _atomic_csv(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    frame.to_csv(temp, index=False, encoding="utf-8")
    os.replace(temp, path)


def _excel_safe(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().replace(tzinfo=None)
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, sort_keys=True)
    return value


def _write_workbook(tables: dict[str, pd.DataFrame], path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    readme = workbook.create_sheet("README")
    readme["A1"] = "Step 9 KPI Read-Only Evaluation V1"
    readme["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    readme["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    readme.merge_cells("A1:F1")
    readme["A3"] = "Status"
    readme["B3"] = STATUS
    readme["A4"] = "Specification"
    readme["B4"] = SPECIFICATION_ID
    readme["A5"] = "Safety"
    readme["B5"] = "Read-only sources; no selector, router, order, morning ledger or EOD outcome is changed."
    readme["A7"] = "Power BI"
    readme["B7"] = "Get Data → Excel workbook → select the named tables. Use Evidence Status as a mandatory slicer."
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 95

    table_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    for table_name, frame in tables.items():
        sheet = workbook.create_sheet(table_name[:31])
        sheet.freeze_panes = "A2"
        values = [list(frame.columns)] + [[_excel_safe(v) for v in row] for row in frame.itertuples(index=False, name=None)]
        if len(values) == 1:
            values.append([None for _ in frame.columns])
        for r_idx, row in enumerate(values, 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(r_idx, c_idx, value)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        last_row = max(2, len(values))
        last_col = max(1, len(frame.columns))
        ref = f"A1:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = table_style
        sheet.add_table(table)
        for idx, col in enumerate(frame.columns, 1):
            sample = [str(x) for x in frame[col].dropna().head(100).tolist()]
            width = min(42, max(10, len(str(col)) + 2, max((len(x) for x in sample), default=0) + 2))
            sheet.column_dimensions[get_column_letter(idx)].width = width

    # Small preview using engine and benchmark daily tables.
    preview = workbook.create_sheet("Dashboard_Preview", 1)
    preview["A1"] = "Step 9 KPI — Daily Overview"
    preview["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    preview["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    preview.merge_cells("A1:H1")
    engine = tables.get("tblEngineDaily", pd.DataFrame())
    benchmark = tables.get("tblBenchmarkDaily", pd.DataFrame())
    data_rows = []
    if not engine.empty:
        for r in engine.loc[engine["run_status"].isin(["COMPLETE", "VALID_NO_TRADE"])].to_dict("records"):
            data_rows.append([r["session_date"], r["evidence_status"], r["engine_book_id"], r["standardized_net_pnl_sek"]])
    if not benchmark.empty:
        for r in benchmark.loc[benchmark["coverage_status"].eq("COMPLETE")].to_dict("records"):
            data_rows.append([r["session_date"], r["evidence_status"], r["benchmark_id"], r["standardized_net_pnl_sek"]])
    preview.append([])
    preview.append(["Session Date", "Evidence Status", "Series", "Daily Standardized P&L SEK"])
    for row in sorted(data_rows, key=lambda x: (str(x[0]), str(x[2]))):
        preview.append(row)
    for cell in preview[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
    preview.column_dimensions["A"].width = 16
    preview.column_dimensions["B"].width = 25
    preview.column_dimensions["C"].width = 42
    preview.column_dimensions["D"].width = 28
    if data_rows:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Daily standardized P&L (use Power BI for multi-day cumulative lines)"
        chart.y_axis.title = "SEK"
        chart.x_axis.title = "Rows / series"
        chart.add_data(Reference(preview, min_col=4, min_row=3, max_row=3 + len(data_rows)), titles_from_data=True)
        chart.set_categories(Reference(preview, min_col=3, min_row=4, max_row=3 + len(data_rows)))
        chart.height = 10
        chart.width = 18
        preview.add_chart(chart, "F3")

    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(dir=path.parent, suffix=".xlsx", delete=False) as handle:
        temp_path = Path(handle.name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        temp_path.unlink(missing_ok=True)


def _output_file_map(output_dir: Path) -> dict[str, Path]:
    return {
        "dimSession": output_dir / "step9kpi_session.csv",
        "dimEngine": output_dir / "step9kpi_engine_dimension.csv",
        "dimStrategy": output_dir / "step9kpi_strategy_dimension.csv",
        "tblEngineDaily": output_dir / "step9kpi_engine_daily.csv",
        "tblBenchmarkDaily": output_dir / "step9kpi_benchmark_daily.csv",
        "tblStrategyOutcome": output_dir / "step9kpi_strategy_outcome.csv",
        "tblStrategyAccuracy": output_dir / "step9kpi_strategy_accuracy.csv",
        "tblRegimeAccuracy": output_dir / "step9kpi_regime_accuracy.csv",
        "tblRegimeStrategyAccuracy": output_dir / "step9kpi_regime_strategy_accuracy.csv",
        "tblRankingTicker": output_dir / "step9kpi_ranking_ticker.csv",
        "tblRankingDaily": output_dir / "step9kpi_ranking_daily.csv",
        "tblPortfolioSize": output_dir / "step9kpi_portfolio_size.csv",
        "tblDataQuality": output_dir / "step9kpi_data_quality.csv",
    }


def build(
    project_root: Path = PROJECT_ROOT,
    config_path: Path | None = None,
    schema_path: Path | None = None,
    output_dir: Path | None = None,
    workbook_path: Path | None = None,
    session_dates: list[str] | None = None,
) -> BuildResult:
    project_root = Path(project_root).resolve()
    config_path = Path(config_path or project_root / "config" / "step9kpi_read_only_evaluation_v1.json")
    schema_path = Path(schema_path or project_root / "config" / "step9kpi_output_schema_v1.json")
    output_dir = Path(output_dir or project_root / "data" / "step9kpi")
    workbook_path = Path(workbook_path or output_dir / "powerbi_step9_kpi_monitor.xlsx")

    config = _read_config(config_path)
    schema = _read_schema(schema_path)
    sources = _build_sources(project_root, config)
    before = _existing_source_hashes(sources)
    batches = _load_engine_batches(sources)
    evidence = _session_evidence_map(batches)
    morning_regime = _morning_regime_map(batches)
    sessions = sorted(session_dates or evidence.keys())
    if not sessions:
        raise SourceContractError("No sealed sessions were found in the source ledgers.")
    evidence = {session: evidence.get(session, "PROSPECTIVE_EXCLUDED") for session in sessions}

    prices = _canonical_prices(sources["price_db"].path)
    regime_tickers, broad_tickers = _load_universes(project_root, config, sources)
    outcomes = _strategy_outcomes(sources, evidence, config)
    if not outcomes.empty:
        outcomes = outcomes.loc[outcomes["session_date"].astype(str).isin(sessions)].reset_index(drop=True)
    strategy_dim = _dim_strategy(project_root, config, outcomes)
    engine_daily = _engine_daily(sources, evidence, config)
    if not engine_daily.empty:
        engine_daily = engine_daily.loc[engine_daily["session_date"].astype(str).isin(sessions)].reset_index(drop=True)
    benchmarks = _benchmarks(outcomes, strategy_dim, evidence)
    strategy_accuracy = _strategy_accuracy(outcomes, strategy_dim, sources, evidence, float(config["pnl_tie_tolerance_sek"]))
    regime_accuracy = _regime_accuracy(prices, regime_tickers, broad_tickers, sessions, morning_regime, evidence, config)
    ranking_ticker, ranking_daily, portfolio_size = _ranking_tables(sources, evidence, config)
    regime_strategy = _regime_strategy_accuracy(engine_daily, outcomes, strategy_dim, regime_accuracy, sources, project_root, config)
    session_dim = _dim_session(sessions, evidence, morning_regime, regime_accuracy, prices)

    raw_tables = {
        "dimSession": session_dim,
        "dimEngine": _dim_engine(),
        "dimStrategy": strategy_dim,
        "tblEngineDaily": engine_daily,
        "tblBenchmarkDaily": benchmarks,
        "tblStrategyOutcome": outcomes,
        "tblStrategyAccuracy": strategy_accuracy,
        "tblRegimeAccuracy": regime_accuracy,
        "tblRegimeStrategyAccuracy": regime_strategy,
        "tblRankingTicker": ranking_ticker,
        "tblRankingDaily": ranking_daily,
        "tblPortfolioSize": portfolio_size,
    }
    tables = {name: _conform_columns(frame, schema, name) for name, frame in raw_tables.items()}
    after = _existing_source_hashes(sources)
    tables["tblDataQuality"] = _conform_columns(
        _data_quality(sessions, evidence, sources, before, after, tables), schema, "tblDataQuality"
    )
    _validate_schema(tables, schema)

    output_files: list[Path] = []
    for table_name, path in _output_file_map(output_dir).items():
        _atomic_csv(tables[table_name], path)
        output_files.append(path)
    _write_workbook(tables, workbook_path)
    output_files.append(workbook_path)

    manifest = {
        "implementation_id": IMPLEMENTATION_ID,
        "specification_id": SPECIFICATION_ID,
        "schema_id": SCHEMA_ID,
        "status": STATUS,
        "generated_at": datetime.now().astimezone().isoformat(),
        "project_root": str(project_root),
        "sessions": sessions,
        "evidence_statuses": evidence,
        "output_files": [str(path) for path in output_files],
        "source_hashes_before": before,
        "source_hashes_after": after,
        "source_files_unchanged": before == after,
        "router_active": False,
        "orders_enabled": False,
    }
    manifest_path = output_dir / "step9kpi_build_manifest.json"
    temp = manifest_path.with_suffix(".json.tmp")
    temp.write_text(json.dumps(manifest, indent=2, sort_keys=True), encoding="utf-8")
    os.replace(temp, manifest_path)
    output_files.append(manifest_path)
    return BuildResult(tables, output_files, before, after)



def _merge_build_results(
    results: list[tuple[str, Path, BuildResult]],
    schema: dict[str, Any],
    output_dir: Path,
    workbook_path: Path,
) -> BuildResult:
    if not results:
        raise SourceContractError("No KPI source projects were successfully built.")

    merged_tables: dict[str, pd.DataFrame] = {}
    for table_name, table_schema in schema["tables"].items():
        frames = [
            result.tables[table_name]
            for _, _, result in results
            if table_name in result.tables and not result.tables[table_name].empty
        ]
        if frames:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", FutureWarning)
                combined = pd.concat(frames, ignore_index=True)
        else:
            combined = pd.DataFrame()
        combined = _conform_columns(combined, schema, table_name)
        primary_key = list(table_schema.get("primary_key", []))
        if not combined.empty and primary_key:
            exact = combined.drop_duplicates().copy()
            duplicate_mask = exact.duplicated(primary_key, keep=False)
            if duplicate_mask.any():
                conflict = exact.loc[duplicate_mask, primary_key].drop_duplicates().head(10).to_dict("records")
                raise OutputContractError(
                    f"Conflicting duplicate rows while merging {table_name}; keys={conflict}"
                )
            combined = exact.drop_duplicates(primary_key, keep="first").reset_index(drop=True)
        merged_tables[table_name] = combined

    _validate_schema(merged_tables, schema)

    output_files: list[Path] = []
    for table_name, path in _output_file_map(output_dir).items():
        _atomic_csv(merged_tables[table_name], path)
        output_files.append(path)
    _write_workbook(merged_tables, workbook_path)
    output_files.append(workbook_path)

    before: dict[str, str] = {}
    after: dict[str, str] = {}
    source_projects: list[dict[str, Any]] = []
    for source_id, root, result in results:
        source_projects.append({
            "source_id": source_id,
            "project_root": str(root),
            "sessions": sorted(
                set(result.tables["dimSession"].get("session_date", pd.Series(dtype=str)).astype(str))
            ),
        })
        for key, value in result.source_hashes_before.items():
            before[f"{source_id}::{key}"] = value
        for key, value in result.source_hashes_after.items():
            after[f"{source_id}::{key}"] = value

    session_rows = merged_tables["dimSession"]
    sessions = sorted(session_rows["session_date"].astype(str).unique()) if not session_rows.empty else []
    evidence_statuses = {
        f"{row['session_date']}::{row['evidence_status']}": row["evidence_status"]
        for row in session_rows[["session_date", "evidence_status"]].drop_duplicates().to_dict("records")
    } if not session_rows.empty else {}

    manifest = {
        "implementation_id": IMPLEMENTATION_ID,
        "specification_id": SPECIFICATION_ID,
        "schema_id": SCHEMA_ID,
        "status": STATUS,
        "generated_at": datetime.now().astimezone().isoformat(),
        "source_projects": source_projects,
        "sessions": sessions,
        "evidence_statuses": evidence_statuses,
        "output_files": [str(path) for path in output_files],
        "source_hashes_before": before,
        "source_hashes_after": after,
        "source_files_unchanged": before == after,
        "router_active": False,
        "orders_enabled": False,
    }
    manifest_path = output_dir / "step9kpi_build_manifest.json"
    temp = manifest_path.with_suffix(".json.tmp")
    temp.write_text(json.dumps(manifest, indent=2, sort_keys=True), encoding="utf-8")
    os.replace(temp, manifest_path)
    output_files.append(manifest_path)
    return BuildResult(merged_tables, output_files, before, after)


def build_with_supplemental_sources(
    project_root: Path = PROJECT_ROOT,
    config_path: Path | None = None,
    schema_path: Path | None = None,
    output_dir: Path | None = None,
    workbook_path: Path | None = None,
    session_dates: list[str] | None = None,
    supplemental_project_roots: list[Path] | None = None,
    supplemental_session_dates: list[str] | None = None,
) -> BuildResult:
    project_root = Path(project_root).resolve()
    config_path = Path(config_path or project_root / "config" / "step9kpi_read_only_evaluation_v1.json").resolve()
    schema_path = Path(schema_path or project_root / "config" / "step9kpi_output_schema_v1.json").resolve()
    output_dir = Path(output_dir or project_root / "data" / "step9kpi").resolve()
    workbook_path = Path(workbook_path or output_dir / "powerbi_step9_kpi_monitor.xlsx").resolve()
    schema = _read_schema(schema_path)

    supplemental_project_roots = [Path(root).resolve() for root in (supplemental_project_roots or [])]
    if not supplemental_project_roots:
        return build(
            project_root=project_root, config_path=config_path, schema_path=schema_path,
            output_dir=output_dir, workbook_path=workbook_path, session_dates=session_dates,
        )

    results: list[tuple[str, Path, BuildResult]] = []
    with tempfile.TemporaryDirectory(prefix="step9kpi_merge_") as temp_dir_text:
        temp_dir = Path(temp_dir_text)
        primary = build(
            project_root=project_root, config_path=config_path, schema_path=schema_path,
            output_dir=temp_dir / "primary", workbook_path=temp_dir / "primary.xlsx",
            session_dates=session_dates,
        )
        results.append(("PRIMARY", project_root, primary))

        for index, supplemental_root in enumerate(supplemental_project_roots, start=1):
            supplemental = build(
                project_root=supplemental_root, config_path=config_path, schema_path=schema_path,
                output_dir=temp_dir / f"supplemental_{index}",
                workbook_path=temp_dir / f"supplemental_{index}.xlsx",
                session_dates=supplemental_session_dates,
            )
            results.append((f"SUPPLEMENTAL_{index}", supplemental_root, supplemental))

        return _merge_build_results(results, schema, output_dir, workbook_path)

def _print_summary(result: BuildResult) -> None:
    print("STEP9KPI_READ_ONLY_EVALUATION_V1: BUILT")
    print("SOURCE FILES BYTE-FOR-BYTE UNCHANGED:", result.source_hashes_before == result.source_hashes_after)
    for name, frame in result.tables.items():
        print(f"{name}: {len(frame)} rows")
    benchmark = result.tables["tblBenchmarkDaily"]
    if not benchmark.empty:
        main = benchmark.loc[benchmark["benchmark_id"].eq("ORACLE_TOP2_OBSERVED_FIXED")]
        for row in main.to_dict("records"):
            print(
                f"ORACLE_TOP2_OBSERVED_FIXED {row['session_date']}: "
                f"{row['selected_tickers']} / {row['standardized_net_pnl_sek']} SEK"
            )
    print("ROUTER ACTIVE: FALSE")
    print("NO ORDER WAS SENT")


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the read-only Step 9 KPI evaluation and Power BI feed.")
    parser.add_argument("command", nargs="?", default="build", choices=["build"])
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT)
    parser.add_argument("--config", type=Path)
    parser.add_argument("--schema", type=Path)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--date", action="append", dest="dates")
    parser.add_argument("--supplemental-project-root", action="append", type=Path, dest="supplemental_project_roots")
    parser.add_argument("--supplemental-date", action="append", dest="supplemental_dates")
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    result = build_with_supplemental_sources(
        project_root=args.project_root,
        config_path=args.config,
        schema_path=args.schema,
        output_dir=args.output_dir,
        workbook_path=args.workbook,
        session_dates=args.dates,
        supplemental_project_roots=args.supplemental_project_roots,
        supplemental_session_dates=args.supplemental_dates,
    )
    _print_summary(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
