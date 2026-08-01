from __future__ import annotations

import argparse
import json
import math
import os
import sqlite3
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from RegimeTrading.scripts import step9q_b_lite_live_trade_feed as step9qb
from RegimeTrading.core.stage_registry import resolve_stage_path


SCHEMA_VERSION = "POWERBI_MONITOR_SCHEMA_V1"
STEP9Q_STATUS = "READ_ONLY_EXCEL_AND_INTRADAY_MONITORING_FEED_NOT_ROUTER_ACTIVE"
STOCKHOLM_TZ = ZoneInfo("Europe/Stockholm")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / "data"
DEFAULT_SCHEMA_PATH = PROJECT_ROOT / "config" / "step9q_powerbi_schema_v1.json"
DEFAULT_OUTPUT_PATH = DATA_DIR / "powerbi" / "powerbi_live_master.xlsx"
DEFAULT_STEP9I_LEDGER = resolve_stage_path("step9i")
DEFAULT_STEP9L_LEDGER = resolve_stage_path("step9l")
DEFAULT_PRICE_DB = resolve_stage_path("prices")

DECISION_BATCH_REQUIRED_COLUMNS = {
    "batch_id",
    "experiment_id",
    "session_date",
    "created_at_stockholm",
    "prospective_status",
    "source_max_datetime",
    "regime_source_tickers_observed",
    "holdout_tickers_observed",
    "primary_regime",
    "regime_confidence",
    "confidence_band",
    "direction_bias",
    "decision_rows",
    "eligible_rows",
    "active_guardrails",
}

DECISION_REQUIRED_COLUMNS = {
    "decision_id",
    "batch_id",
    "session_date",
    "contract_id",
    "test_role",
    "ticker",
    "company_id",
    "broad_sector",
    "primary_regime",
    "regime_match",
    "ticker_relative_state",
    "volatility_bucket",
    "range_state",
    "sector_direction_state",
    "sector_direction_alignment",
    "intended_side",
    "contract_eligible",
    "decision_action",
    "decision_reason",
    "max_router_source_label",
    "point_in_time_pass",
    "sealed_at_stockholm",
}

PRICE_REQUIRED_COLUMNS = {
    "datetime",
    "open",
    "high",
    "low",
    "close",
    "ticker",
}


class Step9QError(RuntimeError):
    """Base error for the read-only Step 9Q-A reporting feed."""


class SourceSchemaError(Step9QError):
    """Raised when an existing source database does not match its frozen schema."""


class WorkbookContractError(Step9QError):
    """Raised when generated reporting tables violate the Excel contract."""


@dataclass(frozen=True)
class EngineSpec:
    engine: str
    ledger_path: Path


@dataclass
class EngineSnapshot:
    spec: EngineSpec
    session_date: str
    batch: pd.DataFrame
    decisions: pd.DataFrame
    status: str


def _stockholm_now() -> datetime:
    return datetime.now(STOCKHOLM_TZ)


def _boolish(value: Any, default: bool = False) -> bool:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off", ""}:
        return False
    return default


def _number(value: Any, default: float | None = None) -> float | None:
    if value is None or pd.isna(value):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _integer(value: Any, default: int = 0) -> int:
    number = _number(value)
    return default if number is None else int(number)


def _parse_local_datetime(value: Any) -> datetime | None:
    if value is None or pd.isna(value) or not str(value).strip():
        return None
    try:
        stamp = pd.Timestamp(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if pd.isna(stamp):
        return None
    if stamp.tzinfo is not None:
        stamp = stamp.tz_convert(STOCKHOLM_TZ).tz_localize(None)
    return stamp.to_pydatetime()


def _parse_utc_datetime(value: Any) -> datetime | None:
    if value is None or pd.isna(value) or not str(value).strip():
        return None
    try:
        stamp = pd.Timestamp(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if pd.isna(stamp):
        return None
    if stamp.tzinfo is not None:
        stamp = stamp.tz_convert("UTC").tz_localize(None)
    return stamp.to_pydatetime()


def _read_schema(path: Path) -> dict[str, Any]:
    if not Path(path).exists():
        raise WorkbookContractError(f"Schema contract does not exist: {path}")
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if payload.get("schema_version") != SCHEMA_VERSION:
        raise WorkbookContractError(
            f"Expected schema {SCHEMA_VERSION}, found {payload.get('schema_version')!r}."
        )
    tables = payload.get("tables")
    if not isinstance(tables, dict) or not tables:
        raise WorkbookContractError("Schema contract contains no reporting tables.")
    return payload


def _sqlite_uri(path: Path) -> str:
    absolute = Path(path).resolve().as_posix()
    return f"file:{quote(absolute, safe='/:')}?mode=ro"


def _connect_read_only(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(_sqlite_uri(path), uri=True)
    connection.execute("PRAGMA query_only = ON")
    return connection


def _table_names(connection: sqlite3.Connection) -> set[str]:
    rows = connection.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()
    return {str(row[0]) for row in rows}


def _table_columns(connection: sqlite3.Connection, table: str) -> set[str]:
    return {str(row[1]) for row in connection.execute(f"PRAGMA table_info({table})")}


def _require_table_schema(
    connection: sqlite3.Connection,
    database_path: Path,
    table: str,
    required_columns: set[str],
) -> None:
    if table not in _table_names(connection):
        raise SourceSchemaError(
            f"Existing source database {database_path} has no table {table!r}."
        )
    columns = _table_columns(connection, table)
    missing = sorted(required_columns - columns)
    if missing:
        raise SourceSchemaError(
            f"Source table {database_path}:{table} is missing columns: "
            + ", ".join(missing)
        )


def _available_sessions(ledger_path: Path) -> list[str]:
    ledger_path = Path(ledger_path)
    if not ledger_path.exists():
        return []
    with closing(_connect_read_only(ledger_path)) as connection:
        _require_table_schema(
            connection,
            ledger_path,
            "shadow_decision_batches",
            DECISION_BATCH_REQUIRED_COLUMNS,
        )
        rows = connection.execute(
            "SELECT session_date FROM shadow_decision_batches ORDER BY session_date"
        ).fetchall()
    return [str(row[0]) for row in rows if row and row[0]]


def _resolve_session_date(
    requested_date: str | None,
    engine_specs: list[EngineSpec],
    now: datetime,
) -> str:
    if requested_date:
        try:
            return date.fromisoformat(requested_date).isoformat()
        except ValueError as exc:
            raise Step9QError("--date must use YYYY-MM-DD format.") from exc

    sessions: list[str] = []
    for spec in engine_specs:
        sessions.extend(_available_sessions(spec.ledger_path))
    return max(sessions) if sessions else now.date().isoformat()


def _read_engine_snapshot(spec: EngineSpec, session_date: str) -> EngineSnapshot:
    path = Path(spec.ledger_path)
    if not path.exists():
        return EngineSnapshot(
            spec=spec,
            session_date=session_date,
            batch=pd.DataFrame(),
            decisions=pd.DataFrame(),
            status="LEDGER_MISSING",
        )

    with closing(_connect_read_only(path)) as connection:
        _require_table_schema(
            connection,
            path,
            "shadow_decision_batches",
            DECISION_BATCH_REQUIRED_COLUMNS,
        )
        _require_table_schema(
            connection,
            path,
            "shadow_decisions",
            DECISION_REQUIRED_COLUMNS,
        )
        batch = pd.read_sql_query(
            "SELECT * FROM shadow_decision_batches WHERE session_date = ?",
            connection,
            params=(session_date,),
        )
        if batch.empty:
            return EngineSnapshot(
                spec=spec,
                session_date=session_date,
                batch=batch,
                decisions=pd.DataFrame(),
                status="NO_BATCH_FOR_SESSION",
            )

        batch_id = str(batch.iloc[0]["batch_id"])
        decisions = pd.read_sql_query(
            """
            SELECT *
            FROM shadow_decisions
            WHERE batch_id = ?
            ORDER BY contract_id, ticker
            """,
            connection,
            params=(batch_id,),
        )

    expected_rows = _integer(batch.iloc[0].get("decision_rows"))
    if decisions.empty:
        status = "BATCH_WITHOUT_DECISIONS"
    elif expected_rows != len(decisions):
        status = "DECISION_ROW_COUNT_MISMATCH"
    else:
        status = "READY"
    return EngineSnapshot(spec, session_date, batch, decisions, status)


def _contract_type(test_role: Any) -> str:
    role = str(test_role or "").strip().upper()
    return {
        "PRIMARY_HYPOTHESIS": "PRIMARY",
        "NEGATIVE_GUARDRAIL": "GUARDRAIL",
        "COMPLEMENT_CONTROL": "CONTROL",
        "EXECUTION_COMPARATOR": "COMPARATOR",
    }.get(role, "OTHER")


def _decision_status(row: pd.Series) -> str:
    action = str(row.get("decision_action", ""))
    eligible = _boolish(row.get("contract_eligible"))
    role = str(row.get("test_role", ""))
    if action == "DATA_INCOMPLETE_NO_SHADOW_DECISION":
        return "DATA_INCOMPLETE"
    if eligible and role == "NEGATIVE_GUARDRAIL":
        return "GUARDRAIL_ACTIVE"
    if eligible:
        return "ELIGIBLE"
    return "INELIGIBLE"


def _build_engine_status(snapshot: EngineSnapshot) -> dict[str, Any]:
    if snapshot.batch.empty:
        return {
            "SessionDate": snapshot.session_date,
            "Engine": snapshot.spec.engine,
            "ExperimentID": None,
            "BatchID": None,
            "SealTimeStockholm": None,
            "ProspectiveStatus": None,
            "PrimaryRegime": None,
            "RegimeConfidence": None,
            "ConfidenceBand": None,
            "DirectionBias": None,
            "DecisionCount": 0,
            "EligibleCount": 0,
            "ActiveGuardrailCount": 0,
            "RegimeSourceTickersObserved": 0,
            "TradableTickersObserved": 0,
            "SourceMaxDateTime": None,
            "EngineStatus": snapshot.status,
        }

    row = snapshot.batch.iloc[0]
    active_guardrails = _integer(row.get("active_guardrails"))
    if not snapshot.decisions.empty:
        active_guardrails = int(
            snapshot.decisions["decision_action"]
            .astype(str)
            .eq("GUARDRAIL_ACTIVE_AVOID_STRATEGY")
            .sum()
        )
    return {
        "SessionDate": snapshot.session_date,
        "Engine": snapshot.spec.engine,
        "ExperimentID": row.get("experiment_id"),
        "BatchID": row.get("batch_id"),
        "SealTimeStockholm": _parse_local_datetime(row.get("created_at_stockholm")),
        "ProspectiveStatus": row.get("prospective_status"),
        "PrimaryRegime": row.get("primary_regime"),
        "RegimeConfidence": _number(row.get("regime_confidence")),
        "ConfidenceBand": row.get("confidence_band"),
        "DirectionBias": row.get("direction_bias"),
        "DecisionCount": len(snapshot.decisions),
        "EligibleCount": int(
            snapshot.decisions.get(
                "contract_eligible", pd.Series(dtype=int)
            ).map(_boolish).sum()
        ),
        "ActiveGuardrailCount": active_guardrails,
        "RegimeSourceTickersObserved": _integer(
            row.get("regime_source_tickers_observed")
        ),
        "TradableTickersObserved": int(
            snapshot.decisions.get("ticker", pd.Series(dtype=str)).nunique()
        ),
        "SourceMaxDateTime": _parse_local_datetime(row.get("source_max_datetime")),
        "EngineStatus": snapshot.status,
    }


def _build_signal_decisions(snapshot: EngineSnapshot) -> list[dict[str, Any]]:
    if snapshot.decisions.empty:
        return [
            {
                "SessionDate": snapshot.session_date,
                "Engine": snapshot.spec.engine,
                "DecisionID": None,
                "BatchID": None,
                "ContractID": None,
                "ContractType": "OTHER",
                "TestRole": None,
                "Ticker": None,
                "CompanyID": None,
                "BroadSector": None,
                "PrimaryRegime": None,
                "RegimeMatch": None,
                "TickerRelativeState": None,
                "VolatilityBucket": None,
                "RangeState": None,
                "SectorDirectionState": None,
                "SectorDirectionAlignment": None,
                "IntendedSide": None,
                "ContractEligible": None,
                "IsPrimary": False,
                "IsGuardrail": False,
                "DecisionAction": None,
                "DecisionReason": None,
                "PointInTimePass": None,
                "MaxRouterSourceLabel": None,
                "SealedAtStockholm": None,
                "DecisionStatus": snapshot.status,
            }
        ]

    rows: list[dict[str, Any]] = []
    for source in snapshot.decisions.to_dict("records"):
        role = str(source.get("test_role", ""))
        rows.append(
            {
                "SessionDate": source.get("session_date", snapshot.session_date),
                "Engine": snapshot.spec.engine,
                "DecisionID": source.get("decision_id"),
                "BatchID": source.get("batch_id"),
                "ContractID": source.get("contract_id"),
                "ContractType": _contract_type(role),
                "TestRole": role,
                "Ticker": source.get("ticker"),
                "CompanyID": source.get("company_id"),
                "BroadSector": source.get("broad_sector"),
                "PrimaryRegime": source.get("primary_regime"),
                "RegimeMatch": _boolish(source.get("regime_match")),
                "TickerRelativeState": source.get("ticker_relative_state"),
                "VolatilityBucket": source.get("volatility_bucket"),
                "RangeState": source.get("range_state"),
                "SectorDirectionState": source.get("sector_direction_state"),
                "SectorDirectionAlignment": source.get(
                    "sector_direction_alignment"
                ),
                "IntendedSide": source.get("intended_side"),
                "ContractEligible": _boolish(source.get("contract_eligible")),
                "IsPrimary": role == "PRIMARY_HYPOTHESIS",
                "IsGuardrail": role == "NEGATIVE_GUARDRAIL",
                "DecisionAction": source.get("decision_action"),
                "DecisionReason": source.get("decision_reason"),
                "PointInTimePass": _boolish(source.get("point_in_time_pass")),
                "MaxRouterSourceLabel": source.get("max_router_source_label"),
                "SealedAtStockholm": _parse_local_datetime(
                    source.get("sealed_at_stockholm")
                ),
                "DecisionStatus": _decision_status(pd.Series(source)),
            }
        )
    return rows


def _build_engine_comparison(
    session_date: str,
    engine_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    by_engine = {row["Engine"]: row for row in engine_rows}
    step9i = by_engine["STEP9I_V2"]
    step9l = by_engine["STEP9L_V3"]
    both_ready = step9i["EngineStatus"] == "READY" and step9l["EngineStatus"] == "READY"
    confidence_i = step9i["RegimeConfidence"]
    confidence_l = step9l["RegimeConfidence"]
    confidence_agreement = bool(
        both_ready
        and confidence_i is not None
        and confidence_l is not None
        and abs(float(confidence_i) - float(confidence_l)) <= 1e-12
    )
    regime_agreement = bool(
        both_ready and step9i["PrimaryRegime"] == step9l["PrimaryRegime"]
    )
    both_prospective = bool(
        both_ready
        and step9i["ProspectiveStatus"] == "PROSPECTIVE_CONFIRMATORY_ELIGIBLE"
        and step9l["ProspectiveStatus"] == "PROSPECTIVE_CONFIRMATORY_ELIGIBLE"
    )
    if not both_ready:
        comparison_status = "INCOMPLETE_ENGINE_INPUT"
    elif not regime_agreement:
        comparison_status = "REGIME_DISAGREEMENT_REVIEW_REQUIRED"
    elif not confidence_agreement:
        comparison_status = "CONFIDENCE_DISAGREEMENT_REVIEW_REQUIRED"
    else:
        comparison_status = "ENGINE_MORNING_CLASSIFICATION_AGREES"

    return {
        "SessionDate": session_date,
        "Step9IStatus": step9i["EngineStatus"],
        "Step9LStatus": step9l["EngineStatus"],
        "RegimeAgreement": regime_agreement,
        "ConfidenceAgreement": confidence_agreement,
        "Step9IRegime": step9i["PrimaryRegime"],
        "Step9LRegime": step9l["PrimaryRegime"],
        "Step9IConfidence": confidence_i,
        "Step9LConfidence": confidence_l,
        "Step9IDecisionCount": step9i["DecisionCount"],
        "Step9LDecisionCount": step9l["DecisionCount"],
        "Step9IEligibleCount": step9i["EligibleCount"],
        "Step9LEligibleCount": step9l["EligibleCount"],
        "EligibleDifference": step9l["EligibleCount"] - step9i["EligibleCount"],
        "Step9IActiveGuardrails": step9i["ActiveGuardrailCount"],
        "Step9LActiveGuardrails": step9l["ActiveGuardrailCount"],
        "BothProspective": both_prospective,
        "ComparisonStatus": comparison_status,
    }


def _read_price_session(price_db: Path, session_date: str) -> pd.DataFrame:
    price_db = Path(price_db)
    if not price_db.exists():
        return pd.DataFrame()
    with closing(_connect_read_only(price_db)) as connection:
        _require_table_schema(
            connection,
            price_db,
            "intraday_prices",
            PRICE_REQUIRED_COLUMNS,
        )
        columns = _table_columns(connection, "intraday_prices")
        collected_expression = (
            "collected_at_utc" if "collected_at_utc" in columns else "'' AS collected_at_utc"
        )
        query = f"""
            SELECT datetime, open, high, low, close, ticker, {collected_expression}
            FROM intraday_prices
            WHERE substr(datetime, 1, 10) = ?
            ORDER BY datetime, ticker
        """
        return pd.read_sql_query(query, connection, params=(session_date,))


def _build_feed_health(
    price_db: Path,
    session_date: str,
    schema: dict[str, Any],
    now: datetime,
    stale_after_minutes: float,
) -> dict[str, Any]:
    universe = schema["universe"]
    regime_tickers = set(universe["regime_source_tickers"])
    tradable_tickers = set(universe["tradable_tickers"])
    expected_universe = regime_tickers | tradable_tickers

    if not Path(price_db).exists():
        return {
            "SessionDate": session_date,
            "LastBarDateTime": None,
            "LastBarLabel": None,
            "ExpectedUniverseTickerCount": len(expected_universe),
            "ReceivedUniverseTickerCount": 0,
            "MissingUniverseTickerCount": len(expected_universe),
            "MissingUniverseTickers": ", ".join(sorted(expected_universe)),
            "ExpectedRegimeSourceTickerCount": len(regime_tickers),
            "ReceivedRegimeSourceTickerCount": 0,
            "MissingRegimeSourceTickerCount": len(regime_tickers),
            "MissingRegimeSourceTickers": ", ".join(sorted(regime_tickers)),
            "ExpectedTradableTickerCount": len(tradable_tickers),
            "ReceivedTradableTickerCount": 0,
            "MissingTradableTickerCount": len(tradable_tickers),
            "MissingTradableTickers": ", ".join(sorted(tradable_tickers)),
            "DataRowsForSession": 0,
            "DuplicateBarCount": 0,
            "InvalidOHLCCount": 0,
            "LatestCollectionUTC": None,
            "DataFreshnessMinutes": None,
            "FeedStatus": "PRICE_DB_MISSING",
            "Notes": "The market-data database does not exist; no source was changed.",
        }

    prices = _read_price_session(price_db, session_date)
    if prices.empty:
        received: set[str] = set()
        last_bar = None
        latest_collection = None
        duplicate_count = 0
        invalid_ohlc = 0
    else:
        prices = prices.copy()
        prices["ticker"] = prices["ticker"].astype(str).str.strip()
        prices["parsed_datetime"] = prices["datetime"].map(_parse_local_datetime)
        received = set(prices["ticker"].dropna().astype(str))
        last_bar = max(
            (value for value in prices["parsed_datetime"] if value is not None),
            default=None,
        )
        latest_collection = max(
            (
                value
                for value in prices["collected_at_utc"].map(_parse_utc_datetime)
                if value is not None
            ),
            default=None,
        )
        duplicate_count = int(
            prices.duplicated(subset=["ticker", "datetime"], keep=False).sum()
        )
        numeric = prices[["open", "high", "low", "close"]].apply(
            pd.to_numeric, errors="coerce"
        )
        invalid_mask = (
            numeric.isna().any(axis=1)
            | numeric.le(0).any(axis=1)
            | numeric["high"].lt(numeric[["open", "low", "close"]].max(axis=1))
            | numeric["low"].gt(numeric[["open", "high", "close"]].min(axis=1))
        )
        invalid_ohlc = int(invalid_mask.sum())

    received_universe = received & expected_universe
    missing_universe = sorted(expected_universe - received)
    missing_regime = sorted(regime_tickers - received)
    missing_tradable = sorted(tradable_tickers - received)

    freshness: float | None = None
    if last_bar is not None:
        now_naive = now.replace(tzinfo=None)
        freshness = max(0.0, (now_naive - last_bar).total_seconds() / 60.0)

    selected_is_today = session_date == now.date().isoformat()
    if prices.empty:
        status = "NO_SESSION_DATA"
        notes = "No market bars were found for the selected session."
    elif invalid_ohlc > 0:
        status = "INVALID_OHLC"
        notes = "One or more market bars failed basic OHLC validation."
    elif missing_regime or missing_tradable:
        status = "MISSING_TICKERS"
        notes = "The session is missing one or more frozen expected tickers."
    elif selected_is_today and freshness is not None and freshness > stale_after_minutes:
        status = "STALE"
        notes = f"Latest bar is older than the {stale_after_minutes:g}-minute threshold."
    elif selected_is_today:
        status = "CURRENT"
        notes = "Current-session bars cover the complete frozen expected universe."
    else:
        status = "HISTORICAL_SESSION_STATIC"
        notes = "Historical session data is static; freshness is informational only."

    return {
        "SessionDate": session_date,
        "LastBarDateTime": last_bar,
        "LastBarLabel": last_bar.strftime("%H:%M") if last_bar else None,
        "ExpectedUniverseTickerCount": len(expected_universe),
        "ReceivedUniverseTickerCount": len(received_universe),
        "MissingUniverseTickerCount": len(missing_universe),
        "MissingUniverseTickers": ", ".join(missing_universe),
        "ExpectedRegimeSourceTickerCount": len(regime_tickers),
        "ReceivedRegimeSourceTickerCount": len(received & regime_tickers),
        "MissingRegimeSourceTickerCount": len(missing_regime),
        "MissingRegimeSourceTickers": ", ".join(missing_regime),
        "ExpectedTradableTickerCount": len(tradable_tickers),
        "ReceivedTradableTickerCount": len(received & tradable_tickers),
        "MissingTradableTickerCount": len(missing_tradable),
        "MissingTradableTickers": ", ".join(missing_tradable),
        "DataRowsForSession": len(prices),
        "DuplicateBarCount": duplicate_count,
        "InvalidOHLCCount": invalid_ohlc,
        "LatestCollectionUTC": latest_collection,
        "DataFreshnessMinutes": freshness,
        "FeedStatus": status,
        "Notes": notes,
    }


def _coerce_contract_value(value: Any, value_type: str, nullable: bool) -> Any:
    is_null = value is None or pd.isna(value)
    if is_null:
        if nullable:
            return None
        return {
            "string": "",
            "integer": 0,
            "float": 0.0,
            "boolean": False,
            "date": date(1900, 1, 1),
            "datetime": datetime(1900, 1, 1),
        }[value_type]

    if value_type == "string":
        return str(value)
    if value_type == "integer":
        return int(float(value))
    if value_type == "float":
        return float(value)
    if value_type == "boolean":
        return _boolish(value)
    if value_type == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value)[:10])
    if value_type == "datetime":
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
        parsed = _parse_local_datetime(value)
        if parsed is None:
            raise WorkbookContractError(f"Invalid datetime value: {value!r}")
        return parsed
    raise WorkbookContractError(f"Unsupported contract type: {value_type}")


def _contract_dataframe(
    logical_name: str,
    rows: list[dict[str, Any]],
    schema: dict[str, Any],
) -> pd.DataFrame:
    table_contract = schema["tables"][logical_name]
    column_specs = table_contract["columns"]
    columns = [item["name"] for item in column_specs]
    frame = pd.DataFrame(rows, columns=columns) if not rows else pd.DataFrame(rows)
    missing = [column for column in columns if column not in frame.columns]
    unexpected = [column for column in frame.columns if column not in columns]
    if missing or unexpected:
        raise WorkbookContractError(
            f"{logical_name} contract mismatch; missing={missing}, unexpected={unexpected}."
        )
    frame = frame[columns].copy()
    for item in column_specs:
        column = item["name"]
        frame[column] = frame[column].map(
            lambda value, spec=item: _coerce_contract_value(
                value, spec["type"], bool(spec.get("nullable", False))
            )
        )
    return frame


def build_reporting_tables(
    *,
    session_date: str,
    snapshots: list[EngineSnapshot],
    price_db: Path,
    schema: dict[str, Any],
    now: datetime,
    stale_after_minutes: float,
) -> dict[str, pd.DataFrame]:
    engine_rows = [_build_engine_status(snapshot) for snapshot in snapshots]
    signal_rows: list[dict[str, Any]] = []
    for snapshot in snapshots:
        signal_rows.extend(_build_signal_decisions(snapshot))

    comparison_row = _build_engine_comparison(session_date, engine_rows)
    feed_row = _build_feed_health(
        price_db,
        session_date,
        schema,
        now,
        stale_after_minutes,
    )
    system_row = {
        "SchemaVersion": SCHEMA_VERSION,
        "GeneratedAtStockholm": now.replace(tzinfo=None),
        "SessionDate": session_date,
        "WorkbookStatus": "VALID",
        "PublishMode": "ATOMIC_REPLACEMENT",
        "SourceAccessMode": "SQLITE_READ_ONLY_QUERY_ONLY",
        "PowerBIRefreshMode": "MANUAL_IMPORT_REFRESH",
        "Step9ILedgerExists": Path(snapshots[0].spec.ledger_path).exists(),
        "Step9LLedgerExists": Path(snapshots[1].spec.ledger_path).exists(),
        "PriceDbExists": Path(price_db).exists(),
        "FeedStatus": feed_row["FeedStatus"],
        "Notes": (
            f"{STEP9Q_STATUS}. Source ledgers and market-data database were read only. "
            "No decision, outcome, strategy, router, or production ORB state was changed."
        ),
    }

    step9qb_rows = step9qb.build_step9qb_rows(
        session_date=session_date,
        step9l_ledger=snapshots[1].spec.ledger_path,
        price_db=price_db,
        now=now,
    )

    raw_tables = {
        "System_Status": [system_row],
        "Engine_Status": engine_rows,
        "Signal_Decisions": signal_rows,
        "Engine_Comparison": [comparison_row],
        "Feed_Health": [feed_row],
        "Live_Trade_Status": step9qb_rows["Live_Trade_Status"],
        "Trade_History": step9qb_rows["Trade_History"],
        "Account_Snapshot": step9qb_rows["Account_Snapshot"],
    }
    return {
        name: _contract_dataframe(name, rows, schema)
        for name, rows in raw_tables.items()
    }


def _worksheet_column_width(column_name: str, values: list[Any]) -> float:
    text_lengths = [len(str(column_name))]
    text_lengths.extend(len(str(value)) for value in values if value is not None)
    width = min(max(text_lengths, default=10) + 2, 42)
    if "Reason" in column_name or column_name in {"Notes", "MissingUniverseTickers"}:
        width = min(max(width, 28), 42)
    if column_name.endswith("DateTime") or "TimeStockholm" in column_name:
        width = max(width, 21)
    return float(max(width, 10))


def _write_workbook(
    output_path: Path,
    tables: dict[str, pd.DataFrame],
    schema: dict[str, Any],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="F4CCCC")
    success_fill = PatternFill("solid", fgColor="D9EAD3")

    for logical_name, table_contract in schema["tables"].items():
        frame = tables[logical_name]
        worksheet = workbook.create_sheet(logical_name)
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            worksheet.append(list(row))

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 24

        max_row = max(worksheet.max_row, 2)
        if worksheet.max_row == 1:
            worksheet.append([None] * len(frame.columns))
            max_row = 2
        max_column = len(frame.columns)
        table_ref = f"A1:{get_column_letter(max_column)}{max_row}"
        excel_table = Table(
            displayName=table_contract["excel_table_name"],
            ref=table_ref,
        )
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(excel_table)

        for column_index, column_spec in enumerate(
            table_contract["columns"], start=1
        ):
            column_name = column_spec["name"]
            values = frame[column_name].tolist()
            worksheet.column_dimensions[get_column_letter(column_index)].width = (
                _worksheet_column_width(column_name, values)
            )
            number_format = "General"
            if column_spec["type"] == "date":
                number_format = "yyyy-mm-dd"
            elif column_spec["type"] == "datetime":
                number_format = "yyyy-mm-dd hh:mm:ss"
            elif column_spec["type"] == "float":
                number_format = "0.0000"
            if column_name in {
                "RegimeConfidence",
                "Step9IConfidence",
                "Step9LConfidence",
                "WinRatePct",
            }:
                number_format = "0.0%"
            elif column_name == "DataFreshnessMinutes":
                number_format = "0.0"
            elif column_name.endswith("SEK"):
                number_format = '#,##0.00 "kr"'
            elif column_name.endswith("Price"):
                number_format = "0.0000"
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = (
                    number_format
                )
                worksheet.cell(row=row_index, column=column_index).alignment = Alignment(
                    vertical="top",
                    wrap_text=column_name in {
                        "Notes",
                        "DecisionReason",
                        "MissingUniverseTickers",
                        "MissingRegimeSourceTickers",
                        "MissingTradableTickers",
                    },
                )

        status_columns = {
            "WorkbookStatus",
            "FeedStatus",
            "EngineStatus",
            "ComparisonStatus",
            "ProspectiveStatus",
            "DecisionStatus",
            "TradeStatus",
            "RecordStatus",
            "ReplayStatus",
        }
        for column_name in status_columns.intersection(frame.columns):
            column_index = frame.columns.get_loc(column_name) + 1
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                text = str(cell.value or "").upper()
                if any(term in text for term in ("MISSING", "MISMATCH", "ERROR", "INVALID", "REQUIRED")):
                    cell.fill = error_fill
                elif any(term in text for term in ("LATE", "STALE", "INCOMPLETE", "NONCONFIRMATORY", "NOT_CONFIRMATORY")):
                    cell.fill = warning_fill
                elif any(term in text for term in ("READY", "VALID", "CURRENT", "AGREES", "ELIGIBLE")):
                    cell.fill = success_fill

    workbook.properties.title = "Step 9Q-B Lite Power BI Excel Monitoring Feed"
    workbook.properties.subject = STEP9Q_STATUS
    workbook.properties.creator = "RegimeTrading Step 9Q-B Lite"
    workbook.save(output_path)


def _validate_workbook(
    workbook_path: Path,
    schema: dict[str, Any],
) -> None:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    expected_sheets = list(schema["tables"])
    if workbook.sheetnames != expected_sheets:
        raise WorkbookContractError(
            f"Workbook sheet mismatch: expected {expected_sheets}, found {workbook.sheetnames}."
        )

    seen_table_names: set[str] = set()
    for logical_name, contract in schema["tables"].items():
        worksheet = workbook[logical_name]
        expected_headers = [column["name"] for column in contract["columns"]]
        actual_headers = [
            worksheet.cell(row=1, column=index).value
            for index in range(1, len(expected_headers) + 1)
        ]
        if actual_headers != expected_headers:
            raise WorkbookContractError(
                f"Header mismatch in {logical_name}: expected {expected_headers}, "
                f"found {actual_headers}."
            )
        tables = list(worksheet.tables.values())
        if len(tables) != 1:
            raise WorkbookContractError(
                f"Expected one named Excel table in {logical_name}, found {len(tables)}."
            )
        display_name = tables[0].displayName
        if display_name != contract["excel_table_name"]:
            raise WorkbookContractError(
                f"Excel table mismatch in {logical_name}: expected "
                f"{contract['excel_table_name']}, found {display_name}."
            )
        if display_name in seen_table_names:
            raise WorkbookContractError(f"Duplicate Excel table name: {display_name}")
        seen_table_names.add(display_name)

    if workbook["System_Status"]["A2"].value != SCHEMA_VERSION:
        raise WorkbookContractError("System_Status does not contain the frozen schema version.")
    workbook.close()


def publish_workbook_atomic(
    output_path: Path,
    tables: dict[str, pd.DataFrame],
    schema: dict[str, Any],
) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f"{output_path.stem}.tmp{output_path.suffix}")
    if temporary_path.exists():
        temporary_path.unlink()
    try:
        _write_workbook(temporary_path, tables, schema)
        _validate_workbook(temporary_path, schema)
        os.replace(temporary_path, output_path)
    except Exception:
        if temporary_path.exists():
            temporary_path.unlink()
        raise
    return "REPLACED_ATOMICALLY"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build the Step 9Q-B Lite fixed-schema Excel feed from Step 9I V2, "
            "Step 9L V3, the shared five-minute market-data database, and read-only intraday replay."
        )
    )
    parser.add_argument("--date", help="Session date in YYYY-MM-DD format.")
    parser.add_argument("--step9i-ledger", type=Path, default=DEFAULT_STEP9I_LEDGER)
    parser.add_argument("--step9l-ledger", type=Path, default=DEFAULT_STEP9L_LEDGER)
    parser.add_argument("--price-db", type=Path, default=DEFAULT_PRICE_DB)
    parser.add_argument("--schema", type=Path, default=DEFAULT_SCHEMA_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--stale-after-minutes", type=float, default=15.0)
    parser.add_argument(
        "--require-both-engines",
        action="store_true",
        help="Fail instead of publishing status rows when either engine lacks a batch.",
    )
    return parser.parse_args()


def run(args: argparse.Namespace) -> dict[str, Any]:
    schema = _read_schema(args.schema)
    now = _stockholm_now()
    engine_specs = [
        EngineSpec("STEP9I_V2", Path(args.step9i_ledger)),
        EngineSpec("STEP9L_V3", Path(args.step9l_ledger)),
    ]
    session_date = _resolve_session_date(args.date, engine_specs, now)
    snapshots = [
        _read_engine_snapshot(spec, session_date) for spec in engine_specs
    ]
    if args.require_both_engines:
        not_ready = [
            f"{snapshot.spec.engine}={snapshot.status}"
            for snapshot in snapshots
            if snapshot.status != "READY"
        ]
        if not_ready:
            raise Step9QError(
                "Both engines are required but are not ready: " + ", ".join(not_ready)
            )

    tables = build_reporting_tables(
        session_date=session_date,
        snapshots=snapshots,
        price_db=Path(args.price_db),
        schema=schema,
        now=now,
        stale_after_minutes=float(args.stale_after_minutes),
    )
    action = publish_workbook_atomic(Path(args.output), tables, schema)

    engine_status = tables["Engine_Status"].set_index("Engine")
    feed = tables["Feed_Health"].iloc[0]
    return {
        "schema_version": SCHEMA_VERSION,
        "session_date": session_date,
        "step9i_rows": int(engine_status.loc["STEP9I_V2", "DecisionCount"]),
        "step9l_rows": int(engine_status.loc["STEP9L_V3", "DecisionCount"]),
        "step9i_eligible": int(engine_status.loc["STEP9I_V2", "EligibleCount"]),
        "step9l_eligible": int(engine_status.loc["STEP9L_V3", "EligibleCount"]),
        "step9l_guardrails": int(
            engine_status.loc["STEP9L_V3", "ActiveGuardrailCount"]
        ),
        "last_market_bar": feed["LastBarDateTime"],
        "feed_status": feed["FeedStatus"],
        "workbook_action": action,
        "output": str(Path(args.output)),
        "live_open_for_trade": int(tables["Live_Trade_Status"]["IsOpenForTrade"].sum()) if not tables["Live_Trade_Status"].empty else 0,
        "live_open_trades": int(tables["Live_Trade_Status"]["IsCurrentlyTraded"].sum()) if not tables["Live_Trade_Status"].empty else 0,
        "live_closed_provisional": int((tables["Live_Trade_Status"]["TradeStatus"].eq("CLOSED_PROVISIONAL") & tables["Live_Trade_Status"]["IsPrimary"]).sum()) if not tables["Live_Trade_Status"].empty else 0,
        "current_equity": float(tables["Account_Snapshot"].iloc[0]["CurrentEquitySEK"]),
        "total_pnl": float(tables["Account_Snapshot"].iloc[0]["TotalPnLSEK"]),
        "replay_status": str(tables["Account_Snapshot"].iloc[0]["ReplayStatus"]),
    }


def main() -> None:
    args = parse_args()
    result = run(args)
    print("\n=== STEP 9Q-B LITE POWER BI EXCEL SNAPSHOT ===")
    print(f"Schema version      : {result['schema_version']}")
    print(f"Session date        : {result['session_date']}")
    print(f"Step 9I rows/eligible: {result['step9i_rows']}/{result['step9i_eligible']}")
    print(f"Step 9L rows/eligible: {result['step9l_rows']}/{result['step9l_eligible']}")
    print(f"Step 9L guardrails  : {result['step9l_guardrails']}")
    print(f"Last market bar     : {result['last_market_bar'] or ''}")
    print(f"Feed status         : {result['feed_status']}")
    print(f"Open for trade      : {result['live_open_for_trade']}")
    print(f"Open trades         : {result['live_open_trades']}")
    print(f"Closed provisional  : {result['live_closed_provisional']}")
    print(f"Current equity      : {result['current_equity']:.2f} SEK")
    print(f"Total PnL           : {result['total_pnl']:.2f} SEK")
    print(f"Replay status       : {result['replay_status']}")
    print("Workbook validation : PASSED")
    print(f"Workbook action     : {result['workbook_action']}")
    print("Source access       : READ_ONLY")
    print(f"Output              : {result['output']}")
    print("No orders were sent and no source ledger was modified.")


if __name__ == "__main__":
    main()
