from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import sqlite3
import sys
import tempfile
import unittest
from contextlib import closing
from datetime import datetime
from pathlib import Path
from unittest.mock import patch
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


MODULE_PATH = (
    Path(__file__).resolve().parents[1]
    / "RegimeTrading"
    / "scripts"
    / "step9q_powerbi_excel_feed.py"
)
SPEC = importlib.util.spec_from_file_location("step9q_feed", MODULE_PATH)
step9q = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = step9q
assert SPEC.loader is not None
SPEC.loader.exec_module(step9q)

SCHEMA_PATH = (
    Path(__file__).resolve().parents[1]
    / "config"
    / "step9q_powerbi_schema_v1.json"
)
SESSION_DATE = "2026-07-27"


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def create_ledger(
    path: Path,
    *,
    engine: str,
    session_date: str = SESSION_DATE,
    eligible_rows: int = 0,
    active_guardrails: int = 0,
) -> None:
    experiment = (
        "PROSPECTIVE_SHADOW_ROUTER_V2_CORE5_PLUS_HOLDOUT18"
        if engine == "STEP9I_V2"
        else "STEP9L_SELECTED_STRATEGY_SHADOW_ENGINE_V3"
    )
    batch_id = f"{engine}-{session_date}"
    decisions = [
        {
            "decision_id": f"{engine}-P",
            "batch_id": batch_id,
            "session_date": session_date,
            "contract_id": "PRIMARY_CONTRACT",
            "test_role": "PRIMARY_HYPOTHESIS",
            "ticker": "ABB.ST",
            "company_id": "ABB",
            "broad_sector": "INDUSTRIALS",
            "primary_regime": "HIGH_VOL_REVERSAL",
            "regime_match": 1,
            "ticker_relative_state": "EARLY_LAGGARD",
            "volatility_bucket": "HIGH_RELATIVE_VOL",
            "range_state": "EXPANDED",
            "sector_direction_state": "UP",
            "sector_direction_alignment": "ALIGNED_WITH_GROUP",
            "intended_side": "LONG",
            "contract_eligible": 1 if eligible_rows >= 1 else 0,
            "decision_action": (
                "ELIGIBLE_FOR_EOD_TRIGGER_EVALUATION"
                if eligible_rows >= 1
                else "INELIGIBLE_STATE_FILTER"
            ),
            "decision_reason": "test primary",
            "max_router_source_label": "09:40",
            "point_in_time_pass": 1,
            "sealed_at_stockholm": f"{session_date} 09:46:00+0200",
        },
        {
            "decision_id": f"{engine}-G",
            "batch_id": batch_id,
            "session_date": session_date,
            "contract_id": "GUARDRAIL_CONTRACT",
            "test_role": "NEGATIVE_GUARDRAIL",
            "ticker": "VOLV-B.ST",
            "company_id": "VOLVO_GROUP",
            "broad_sector": "INDUSTRIALS",
            "primary_regime": "HIGH_VOL_REVERSAL",
            "regime_match": 1,
            "ticker_relative_state": "EARLY_LEADER",
            "volatility_bucket": "MEDIUM_RELATIVE_VOL",
            "range_state": "NORMAL",
            "sector_direction_state": "UP",
            "sector_direction_alignment": "ALIGNED_WITH_GROUP",
            "intended_side": "LONG",
            "contract_eligible": 1 if active_guardrails >= 1 else 0,
            "decision_action": (
                "GUARDRAIL_ACTIVE_AVOID_STRATEGY"
                if active_guardrails >= 1
                else "INELIGIBLE_REGIME_MISMATCH"
            ),
            "decision_reason": "test guardrail",
            "max_router_source_label": "09:40",
            "point_in_time_pass": 1,
            "sealed_at_stockholm": f"{session_date} 09:46:00+0200",
        },
    ]

    with closing(sqlite3.connect(path)) as con:
        con.execute(
            """
            CREATE TABLE shadow_decision_batches (
                batch_id TEXT,
                experiment_id TEXT,
                session_date TEXT,
                created_at_stockholm TEXT,
                prospective_status TEXT,
                source_max_datetime TEXT,
                regime_source_tickers_observed INTEGER,
                holdout_tickers_observed INTEGER,
                primary_regime TEXT,
                regime_confidence REAL,
                confidence_band TEXT,
                direction_bias TEXT,
                decision_rows INTEGER,
                eligible_rows INTEGER,
                active_guardrails INTEGER,
                taxonomy_payload_json TEXT,
                research_max_concurrent_ideas INTEGER
            )
            """
        )
        con.execute(
            """
            CREATE TABLE shadow_decisions (
                decision_id TEXT,
                batch_id TEXT,
                session_date TEXT,
                contract_id TEXT,
                test_role TEXT,
                ticker TEXT,
                company_id TEXT,
                broad_sector TEXT,
                primary_regime TEXT,
                regime_match INTEGER,
                ticker_relative_state TEXT,
                volatility_bucket TEXT,
                range_state TEXT,
                sector_direction_state TEXT,
                sector_direction_alignment TEXT,
                intended_side TEXT,
                contract_eligible INTEGER,
                decision_action TEXT,
                decision_reason TEXT,
                max_router_source_label TEXT,
                point_in_time_pass INTEGER,
                sealed_at_stockholm TEXT
            )
            """
        )
        con.execute(
            """
            INSERT INTO shadow_decision_batches VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
            """,
            (
                batch_id,
                experiment,
                session_date,
                f"{session_date} 09:46:00+0200",
                "PROSPECTIVE_CONFIRMATORY_ELIGIBLE",
                f"{session_date} 09:40:00",
                11,
                23,
                "HIGH_VOL_REVERSAL",
                0.899,
                "HIGH",
                "MIXED",
                len(decisions),
                eligible_rows + active_guardrails,
                active_guardrails,
                '{"date":"' + session_date + '","primary_regime":"HIGH_VOL_REVERSAL","direction_bias":"NEUTRAL","research_risk_multiplier":0.5,"research_max_concurrent_ideas":2}',
                2,
            ),
        )
        columns = list(decisions[0])
        placeholders = ",".join("?" for _ in columns)
        con.executemany(
            f"INSERT INTO shadow_decisions ({','.join(columns)}) VALUES ({placeholders})",
            [[row[column] for column in columns] for row in decisions],
        )
        con.commit()


def create_price_db(path: Path, missing_ticker: str | None = None) -> None:
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    expected = sorted(
        set(schema["universe"]["regime_source_tickers"])
        | set(schema["universe"]["tradable_tickers"])
    )
    if missing_ticker:
        expected.remove(missing_ticker)
    with closing(sqlite3.connect(path)) as con:
        con.execute(
            """
            CREATE TABLE intraday_prices (
                datetime TEXT,
                open REAL,
                high REAL,
                low REAL,
                close REAL,
                ticker TEXT,
                source TEXT,
                collected_at_utc TEXT
            )
            """
        )
        con.executemany(
            "INSERT INTO intraday_prices VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            [
                (
                    f"{SESSION_DATE} 09:40:00",
                    100.0,
                    101.0,
                    99.0,
                    100.5,
                    ticker,
                    "TEST",
                    f"{SESSION_DATE} 07:46:00",
                )
                for ticker in expected
            ],
        )
        con.commit()


def args_for(
    root: Path,
    *,
    require_both: bool = False,
    step9l_exists: bool = True,
) -> argparse.Namespace:
    step9i = root / "step9i.db"
    step9l = root / "step9l.db"
    prices = root / "prices.db"
    output = root / "powerbi_live_master.xlsx"
    create_ledger(step9i, engine="STEP9I_V2", eligible_rows=0)
    if step9l_exists:
        create_ledger(
            step9l,
            engine="STEP9L_V3",
            eligible_rows=1,
            active_guardrails=1,
        )
    create_price_db(prices)
    return argparse.Namespace(
        date=SESSION_DATE,
        step9i_ledger=step9i,
        step9l_ledger=step9l,
        price_db=prices,
        schema=SCHEMA_PATH,
        output=output,
        stale_after_minutes=15.0,
        require_both_engines=require_both,
    )


class Step9QPowerBIExcelFeedTests(unittest.TestCase):
    def test_builds_named_tables_and_never_changes_source_databases(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            args = args_for(root)
            before = {
                path.name: file_hash(path)
                for path in (args.step9i_ledger, args.step9l_ledger, args.price_db)
            }

            result = step9q.run(args)

            self.assertEqual(result["step9i_rows"], 2)
            self.assertEqual(result["step9l_rows"], 2)
            self.assertEqual(result["step9l_eligible"], 2)
            self.assertTrue(args.output.exists())
            after = {
                path.name: file_hash(path)
                for path in (args.step9i_ledger, args.step9l_ledger, args.price_db)
            }
            self.assertEqual(before, after)

            workbook = load_workbook(args.output)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "System_Status",
                    "Engine_Status",
                    "Signal_Decisions",
                    "Engine_Comparison",
                    "Feed_Health",
                    "Live_Trade_Status",
                    "Trade_History",
                    "Account_Snapshot",
                ],
            )
            expected_tables = {
                "System_Status": "tblSystemStatus",
                "Engine_Status": "tblEngineStatus",
                "Signal_Decisions": "tblSignalDecisions",
                "Engine_Comparison": "tblEngineComparison",
                "Feed_Health": "tblFeedHealth",
                "Live_Trade_Status": "tblLiveTradeStatus",
                "Trade_History": "tblTradeHistory",
                "Account_Snapshot": "tblAccountSnapshot",
            }
            for sheet_name, table_name in expected_tables.items():
                tables = list(workbook[sheet_name].tables.values())
                self.assertEqual(len(tables), 1)
                self.assertEqual(tables[0].displayName, table_name)
            workbook.close()

    def test_primary_and_guardrail_rows_remain_separate(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            args = args_for(Path(directory))
            step9q.run(args)
            workbook = load_workbook(args.output, data_only=True)
            worksheet = workbook["Signal_Decisions"]
            headers = [cell.value for cell in worksheet[1]]
            rows = [dict(zip(headers, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
            v3_rows = [row for row in rows if row["Engine"] == "STEP9L_V3"]
            primary = next(row for row in v3_rows if row["TestRole"] == "PRIMARY_HYPOTHESIS")
            guardrail = next(row for row in v3_rows if row["TestRole"] == "NEGATIVE_GUARDRAIL")
            self.assertEqual(primary["ContractType"], "PRIMARY")
            self.assertTrue(primary["IsPrimary"])
            self.assertFalse(primary["IsGuardrail"])
            self.assertEqual(guardrail["ContractType"], "GUARDRAIL")
            self.assertFalse(guardrail["IsPrimary"])
            self.assertTrue(guardrail["IsGuardrail"])
            self.assertEqual(guardrail["DecisionStatus"], "GUARDRAIL_ACTIVE")
            workbook.close()

    def test_repeated_export_replaces_snapshot_without_duplication(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            args = args_for(Path(directory))
            step9q.run(args)
            first = load_workbook(args.output)
            first_rows = first["Signal_Decisions"].max_row
            first.close()
            step9q.run(args)
            second = load_workbook(args.output)
            second_rows = second["Signal_Decisions"].max_row
            second.close()
            self.assertEqual(first_rows, second_rows)

    def test_failed_validation_preserves_previous_valid_output(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            args = args_for(root)
            schema = step9q._read_schema(args.schema)
            now = datetime(2026, 7, 27, 9, 47, tzinfo=ZoneInfo("Europe/Stockholm"))
            specs = [
                step9q.EngineSpec("STEP9I_V2", args.step9i_ledger),
                step9q.EngineSpec("STEP9L_V3", args.step9l_ledger),
            ]
            snapshots = [step9q._read_engine_snapshot(spec, SESSION_DATE) for spec in specs]
            tables = step9q.build_reporting_tables(
                session_date=SESSION_DATE,
                snapshots=snapshots,
                price_db=args.price_db,
                schema=schema,
                now=now,
                stale_after_minutes=15.0,
            )
            args.output.write_bytes(b"PREVIOUS_VALID_WORKBOOK")
            with patch.object(
                step9q,
                "_validate_workbook",
                side_effect=step9q.WorkbookContractError("forced failure"),
            ):
                with self.assertRaises(step9q.WorkbookContractError):
                    step9q.publish_workbook_atomic(args.output, tables, schema)
            self.assertEqual(args.output.read_bytes(), b"PREVIOUS_VALID_WORKBOOK")

    def test_missing_engine_is_visible_and_strict_mode_fails(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            args = args_for(root, step9l_exists=False)
            step9q.run(args)
            workbook = load_workbook(args.output, data_only=True)
            worksheet = workbook["Engine_Status"]
            headers = [cell.value for cell in worksheet[1]]
            rows = [dict(zip(headers, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
            step9l = next(row for row in rows if row["Engine"] == "STEP9L_V3")
            self.assertEqual(step9l["EngineStatus"], "LEDGER_MISSING")
            workbook.close()

            strict_root = root / "strict"
            strict_root.mkdir()
            strict_args = args_for(strict_root, require_both=True, step9l_exists=False)
            with self.assertRaises(step9q.Step9QError):
                step9q.run(strict_args)
            self.assertFalse(strict_args.output.exists())

    def test_feed_health_flags_a_missing_frozen_ticker(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            args = args_for(root)
            args.price_db.unlink()
            create_price_db(args.price_db, missing_ticker="ABB.ST")
            step9q.run(args)
            workbook = load_workbook(args.output, data_only=True)
            worksheet = workbook["Feed_Health"]
            headers = [cell.value for cell in worksheet[1]]
            row = dict(zip(headers, next(worksheet.iter_rows(min_row=2, values_only=True))))
            self.assertEqual(row["FeedStatus"], "MISSING_TICKERS")
            self.assertIn("ABB.ST", row["MissingTradableTickers"])
            workbook.close()

    def test_auto_session_selection_uses_latest_available_batch(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            args = args_for(root)
            args.date = None
            with closing(sqlite3.connect(args.step9i_ledger)) as con:
                con.execute(
                    """
                    INSERT INTO shadow_decision_batches VALUES (
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                    )
                    """,
                    (
                        "LATER",
                        "TEST",
                        "2026-07-28",
                        "2026-07-28 09:46:00+0200",
                        "PROSPECTIVE_CONFIRMATORY_ELIGIBLE",
                        "2026-07-28 09:40:00",
                        11,
                        23,
                        "TREND_UP",
                        0.8,
                        "HIGH",
                        "UP",
                        0,
                        0,
                        0,
                        '{"date":"2026-07-28","primary_regime":"TREND_UP","direction_bias":"UP","research_risk_multiplier":0.5,"research_max_concurrent_ideas":2}',
                        2,
                    ),
                )
                con.commit()
            result = step9q.run(args)
            self.assertEqual(result["session_date"], "2026-07-28")


if __name__ == "__main__":
    unittest.main()
